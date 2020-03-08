from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.core.mail import EmailMessage
from celery import task
from celery.task.schedules import crontab
from celery.decorators import periodic_task
import json
from zap_apps.order.models import Order, Transaction, Return, OrderTracker
from zap_apps.address.models import Address
from zap_apps.zap_catalogue.models import ProductImage, ApprovedProduct
from zap_apps.logistics.models import DelhiveryPincode, AramexStatus
from zap_apps.order.order_serializer import ReturnSerializer, TransactionSerializer
from zap_apps.logistics.models import Logistics, LogisticsLog
from zap_apps.payment.payment_serializer import PayoutSerializer
from zap_apps.logistics.logistics_serializer import LogisticsSerializer, LogisticsLogSerializer
from zap_apps.payment.models import PaymentResponse
from django.conf import settings
from django.utils import timezone
import requests
import re
import inflect
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from httplib2 import Http
from urllib import urlencode
from Crypto.Cipher import DES
import pdfkit
# from xhtml2pdf import pisa
import urllib2
import urllib
from django.template import Context
from django.template.loader import render_to_string
import locale
import os
import base64
from zap_apps.order.order_serializer import OrderSerializer
from django.db.models import Q
#DELHIVERY_API_KEY = "iBkrW2RHlFC7WpQtZqz4Dg"
from django.conf import settings
# from zapyle_new.settings.common import DELHIVERY_API_KEY
# DELHIVERY_API_KEY = "187f3b3ae650cf1bda46de6015913b86da7349cc"

from zap_apps.zap_notification.views import ZapSms, ZapEmail
import pdb
from itertools import chain
from zap_apps.zap_notification.tasks import zaplogging
from datetime import datetime, time
# from django.db.models.query import QuerySet
# from djcelery.models import PeriodicTask


# def send_email_task(msg):
#     msg.send()


# #@periodic_task(run_every=(crontab(minute=0, hour=2)))
"""
Increases the quantity after delivery of the returned product according to size
of the returned product.
"""
@task
def increase_quantity_with_size(size_str, product_id, qty):
    if size_str.lower() == 'freesize':
        size = Size.objects.get(size='FS')

    else:
        size_list = re.split('_| |-', size_str)
        if size_list[0].lower() == 'us':
            size = Size.objects.get(us_size=size_list[1])
        elif size_list[0].lower() == 'uk':
            size = Size.objects.get(uk_size=size_list[1])
        else:
            size = Size.objects.get(eu_size=size_list[1])
    pro = NumberOfProducts.objects.get(size=size, product=product_id)
    pro.quantity += qty
    pro.save()

"""
Not in use anymore
"""
def parcelled_pickup_check(pincodes):
    if not isinstance(pincodes, list):
        pincodes = list(pincodes)
    pincodes_str = ",".join(filter(None, pincodes))
    url = settings.PARCELLED_BASE_URL + 'pickup_serviceability/' + pincodes_str
    headers = {'parcelled-api-key': settings.PARCELLED_API_KEY,
               'Content-Type': 'application/json'}
    resp = requests.get(url, headers=headers)
    json_response = resp.json()
    return json_response

"""
Works on top of Order model. Checks first for bangalore and assigns ZAPYLE DELIVERY(Amjad), 
and then the priority is aramex -> delhivery.
"""
# @periodic_task(run_every=(crontab(minute=30, hour=[1, 4])))
def logistics_evaluator(order_number=[], ret_ids=[]):
    # pdb.set_trace()
    orders_list, orders_id_list, all_pincode = [], [], []
    if order_number:
        orders_list = Order.objects.filter(id__in=order_number,
                                           order_status='confirmed')
    else:
        orders_list = Order.objects.filter(
            order_status='confirmed',id=0)
    returns_queryset, all_return_delhivery_object, all_return_pincode = [], [], []
    if ret_ids:
        returns_queryset = Return.objects.filter(id__in=ret_ids,
                                                 return_status='confirmed', approved=True, self_return=False)

    else:
        returns_queryset = Return.objects.filter(
            order_id__order_status='return_requested', approved=True, self_return=False,id=0)
    if orders_list:
        print orders_list,'ordrs_list 643'
        all_pincode = orders_list.values_list('consignor__pincode', flat=True)
    # print returns_queryset, 'returns_queryset'
    if returns_queryset:
        all_return_pincode = returns_queryset.values_list(
            'consignor__pincode', flat=True)
        all_pincode = list(chain(all_pincode, all_return_pincode))
        all_return_delhivery_object = DelhiveryPincode.objects.filter(
            pincode__in=all_return_pincode).values_list('pincode', 'repl')
        pincode_finder = dict(all_return_delhivery_object)

    # if all_pincode:
    #     parcelled_json_resp = parcelled_pickup_check(
    #         all_pincode)
    #     if parcelled_json_resp["meta"]["code"] == 200:
    #         parcelled_pickup_finder = parcelled_json_resp["data"]
    #     else:
    #         parcelled_pickup_finder = {}

    evaluated_order_dict = {}
    for order in orders_list:

        # evaluated_dict = {}
        if order.ordered_product.with_zapyle:
            order.product_verification = 'approved'
            # order.order_status = 'verification'
            order.save()
            pickup_logistics = None
        elif 'bangalore' in order.consignor.city.lower() or 'bengaluru' in order.consignor.city.lower():
            pickup_logistics = 'ZP'
        elif settings.ARAMEX_SERVICE:
            pickup_logistics = 'AR'
        # elif parcelled_pickup_finder.get(order.consignor.pincode, False) and settings.PARCELLED_SERVICE:
        #     pickup_logistics = 'PR'
        elif settings.DELHIVERY_SERVICE:
            pickup_logistics = 'DL'

        if 'bangalore' in order.transaction.consignee.city.lower() or 'bengaluru' in order.transaction.consignee.city.lower():
            delivery_logistics = 'ZP'
        elif settings.ARAMEX_SERVICE:
            delivery_logistics = 'AR'
        elif settings.DELHIVERY_SERVICE:
            delivery_logistics = 'DL'

        evaluated_dict = {order.id: {
            'pickup_logistics': pickup_logistics, 'delivery_logistics': delivery_logistics}}
        evaluated_order_dict.update(evaluated_dict)

    evaluated_returns_dict = {}
    for returns in returns_queryset:

        # Do evaluation of self return
        repl = pincode_finder.get(returns.consignor.pincode, 'N')
        if repl == 'N':
            # returns_id_list.remove(returns.id)
            returns_serializer = ReturnSerializer(data={'self_return': True})
            if returns_serializer.is_valid():
                returns_serializer.update(
                    returns, returns_serializer.validated_data)
                pickup_logistics = 'SR'

        else:
            if returns.consignor.city.lower() == 'bangalore' or returns.consignor.city.lower() == 'bengaluru':
                pickup_logistics = 'ZP'
            elif settings.ARAMEX_SERVICE:
                pickup_logistics = 'AR'
            # elif parcelled_pickup_finder.get(returns.consignor.pincode, False) and settings.PARCELLED_SERVICE:
            #     pickup_logistics = 'PR'
            elif settings.DELHIVERY_SERVICE:
                pickup_logistics = 'DL'

        if returns.order_id.ordered_product.with_zapyle:
            delivery_logistics = None
        elif returns.consignee.city.lower() == 'bangalore' or returns.consignee.city.lower() == 'bengaluru':
            delivery_logistics = 'ZP'
        elif settings.ARAMEX_SERVICE:
            delivery_logistics = 'AR'
        elif settings.DELHIVERY_SERVICE:
            delivery_logistics = 'DL'
        returns_dict = {returns.id: {
            'pickup_logistics': pickup_logistics, 'delivery_logistics': delivery_logistics}}
        evaluated_returns_dict.update(returns_dict)
    # pdb.set_trace()
    if evaluated_order_dict:
        print 'call optimize_order_logistics',evaluated_order_dict
        if 0 and settings.CELERY_USE:
            optimize_order_logistics.delay(evaluated_order_dict)
        else:
            optimize_order_logistics(
                evaluated_order_dict)
        print 'end optimize_order_logistics'
        return evaluated_order_dict
    if evaluated_returns_dict:
    # if settings.CELERY_USE:
        print 'call optimize_return_logistics'
        if 0 and ret_ids:
            optimize_return_logistics.delay(evaluated_returns_dict)
        else:
            optimize_return_logistics(
                evaluated_returns_dict)
        print 'end optimize_return_logistics'
        return evaluated_returns_dict

"""
Clubs returns into 1 logistics if buyer and seller of different orders are same.
"""
@task
def optimize_return_logistics(evaluated_dict):

    returns_object_list = Return.objects.filter(id__in=evaluated_dict.keys())
    returns_list = list(returns_object_list)

    out_index = 0
    for returns in returns_list:
        logistics_returns_id = [returns.id]
        in_index = 0
        if evaluated_dict.get(returns.id).get("pickup_logistics") == 'SR':
            data = {'returns': logistics_returns_id, 'consignee': returns.consignee.id, 'consignor': returns.consignor.id,
                    'triggered_pickup': False, 'pickup_partner': evaluated_dict.get(returns.id).get("pickup_logistics", None), 'product_delivery_partner': evaluated_dict.get(returns.id).get("delivery_logistics", None)}
            log_serializer = LogisticsSerializer(data=data)
            if log_serializer.is_valid():
                log_item = log_serializer.save()
                returns.return_status = 'confirmed'
                returns.save()

            # order = returns.order_id
            # order.order_status = 'return_in_process'
            # order.save()

            logistics_log_data = {'logistics': log_item.id, 'waybill': 'SELF_' + str(returns.id), 'log_status': 1, 'track': True, 'returns': True,
                                  'whole_response': 'Track the self return manually.Follow up with buyer', 'logistics_ref': str(log_item.id), 'partner': 'SR'}
            # print 'success11'
            pickup_log_serializer = LogisticsLogSerializer(
                data=logistics_log_data)
            if pickup_log_serializer.is_valid():
                pickup_log_serializer.save()

        else:
            # order = returns.order_id
            # order.order_status = 'return_in_process'
            # order.save()

            for same_returns in returns_list[out_index + 1:]:
                if ((same_returns.consignee == returns.consignee) and (same_returns.consignor == returns.consignor)):

                    # order = same_returns.order_id
                    # order.order_status = 'return_in_process'
                    # order.save()

                    logistics_returns_id.append(same_returns.id)
                    returns_list.pop(in_index + out_index + 1)
                    del evaluated_dict[same_order.id]
                    in_index -= 1
                in_index += 1

            # CHECK SR

            data = {'returns': logistics_returns_id, 'consignee': returns.consignee.id, 'consignor': returns.consignor.id,
                    'triggered_pickup': False, 'pickup_partner': evaluated_dict.get(returns.id).get("pickup_logistics", None), 'product_delivery_partner': evaluated_dict.get(returns.id).get("delivery_logistics", None)}
            log_serializer = LogisticsSerializer(data=data)
            if log_serializer.is_valid():
                log_serializer.save()
                returns.return_status = 'confirmed'
                returns.save()

        out_index += 1

"""
Clubs orders into 1 logistics if buyer and seller of different orders are same.
"""
@task
def optimize_order_logistics(evaluated_dict):
    print evaluated_dict,'optimize_order_logistics(param)'
    # pdb.set_trace()
    orders_list = Order.objects.filter(id__in=evaluated_dict.keys())
    orders = list(orders_list)
    print orders,'7888888888888'
    out_index = 0
    for order in orders:
        logistics_order_id = [order.id]
        in_index = 0
        OrderTracker.objects.create(orders_id=order.id,status="pickup_in_process")
        if order.ordered_product.with_zapyle:
            order.order_status = 'verification'
            OrderTracker.objects.create(orders_id=order.id,status="picked_up")
            OrderTracker.objects.create(orders_id=order.id,status="verification")
            OrderTracker.objects.create(orders_id=order.id,status="product_approved")

        else:
            order.order_status = 'pickup_in_process'
        order.save()

        # if not :
        for same_order in orders[out_index + 1:]:
            if (((same_order.transaction.consignee == order.transaction.consignee) and (same_order.consignor == order.consignor)
                 and not (same_order.ordered_product.with_zapyle or order.ordered_product.with_zapyle))
                    or ((same_order.ordered_product.with_zapyle and order.ordered_product.with_zapyle) and (same_order.transaction.consignee == order.transaction.consignee))):
                if same_order.ordered_product.with_zapyle:
                    same_order.order_status = 'product_approved'
                else:
                    same_order.order_status = 'pickup_in_process'
                same_order.save()

                logistics_order_id.append(same_order.id)
                orders.pop(in_index + out_index + 1)
                del evaluated_dict[same_order.id]
                in_index -= 1
            in_index += 1
            # evaluated_dict.get(order.id).get("pickup_logistics", None)
        data = {'orders': logistics_order_id, 'consignee': order.transaction.consignee.id, 'consignor': order.consignor.id, 'pickup_partner': evaluated_dict.get(order.id).get(
            "pickup_logistics", None), 'product_delivery_partner': evaluated_dict.get(order.id).get("delivery_logistics", None)}
        log_serializer = LogisticsSerializer(data=data)
        if log_serializer.is_valid():
            log_serializer.save()
        else:
            print log_serializer.errors,' errors 330'

        out_index += 1

"""
Creates a Delhivery waybill.
"""
def delhivery_create_waybill():
    waybil_url = settings.DELHIVERY_BASE_URL + "/waybill/api/fetch/json/?token=" + \
        settings.DELHIVERY_API_KEY + "&cl=" + settings.DELHIVERY_PICKUP_NAME
    headers = {'Accept': 'application/json',
               'Authorization': 'Token ' + settings.DELHIVERY_API_KEY}
    # print requests.get(waybil_url)
    waybill_resp = requests.get(waybil_url, headers=headers)
    waybill_json = waybill_resp.json()
    return waybill_json

"""
Creating Delhivery shipment.
Delhivery API call.
"""
def delhivery_create(packaging_shipments, seller_add, waybill_json):
    # pdb.set_trace()
    packaging_data = json.dumps(
        {'shipments': [packaging_shipments], 'pickup_location': seller_add})
    # data=json.dumps(packaging_data)
    url = settings.DELHIVERY_BASE_URL + \
        '/cmu/push/json/?token=' + settings.DELHIVERY_API_KEY

    data = {'format': 'json', 'data': packaging_data}
    try:
        print 'test1'
        h = Http()
        print 'test2'
        # headers = {u'content-type': u'application/x-www-form-urlencoded'}
        # resp = requests.post(url, params=, headers=headers)
        # packaging_json_response = resp.json()
        msg = "delhivery-create-data-{}".format(json.dumps(data))
        zaplogging.delay(msg) if settings.CELERY_USE else zaplogging(msg)
        resp, content = h.request(url, "POST", urlencode(data))

        msg = "delhivery-create-response-{}".format(content)
        zaplogging.delay(msg) if settings.CELERY_USE else zaplogging(msg)
        json_response = json.loads(content)
        # if 'success' in packaging_json_response and
        # packaging_json_response['success'] is True:
        return json_response, content
        # return None
    except:
        return json_response

"""
Returns packing slip details.
"""
def delhivery_packing_slip(waybill_str):
    url_pm = settings.DELHIVERY_BASE_URL + \
        '/api/p/packing_slip?wbns=' + waybill_str
    print 'code 2022'
    try:
        headers = {'Accept': 'application/json',
                   'Authorization': 'Token ' + settings.DELHIVERY_API_KEY}
        msg = "delhivery-packing-slip-url-{}".format(url_pm)
        zaplogging.delay(msg) if settings.CELERY_USE else zaplogging(msg)
        resp = requests.get(url_pm, headers=headers)
        msg = "delhivery-packing-slip-response-{}".format(str(resp))
        zaplogging.delay(msg) if settings.CELERY_USE else zaplogging(msg)
        json_response = resp.json()
        print json_response
        return json_response
    except:
        return None

"""
Obsolete
"""
def parcelled_create(logistic_id, booking_items, consignee, consignor):
    # ######pdb.set_trace()
    parent_order = {'order_id': logistic_id}

    url = settings.PARCELLED_BASE_URL + 'bookings/'
    headers = {'parcelled-api-key': settings.PARCELLED_API_KEY,
               'Content-Type': 'application/json'}
    data = {'booking_items': booking_items, 'consignee': consignee,
            'consignor': consignor, 'parent_order': parent_order}
    try:
        resp = requests.post(url, data=json.dumps(data), headers=headers)
        json_response = resp.json()
        return json_response
    except Exception as e:
        print e
        return None

"""
Aramex API call for creating pickup.
"""
def aramex_pickup_create(logistic_ref, qty, pickup_address, pickup_contact, payment_type, extra, ret=""):

    url = settings.ARAMEX_BASE_URL + settings.ARAMEX_CREATE_PICKUP_URL

    client_info = {
        "UserName": settings.ARAMEX_USERNAME,
        "Password": settings.ARAMEX_PASSWORD,
        "Version": settings.ARAMEX_VERSION,
        "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
        "AccountPin": settings.ARAMEX_ACCOUNT_PIN,
        "AccountEntity": settings.ARAMEX_ACCOUNT_ENTITY,
        "AccountCountryCode": settings.ARAMEX_ACCOUNT_COUNTRY_CODE
    }
    label_info = {
        "ReportID": 9201,
        "ReportType": "URL"
    }
    pickup = {
        "PickupAddress": pickup_address,
        "PickupContact": pickup_contact,
        "PickupLocation": "Reception",

        "PickupDate": "/Date(" + extra['aramex_pickup_datetime'] + "+0530)/",
        "ReadyTime": "/Date(" + extra['aramex_pickup_readytime'] + "+0530)/",
        "LastPickupTime": "/Date(" + extra['aramex_pickup_lasttime'] + "+0530)/",
        "ClosingTime": "/Date(" + extra['aramex_pickup_closetime'] + "+0530)/",

        "Comments": "Coordinate with the consignor.",
        "Reference1": logistic_ref,
        "Vehicle": "",
        "PickupItems": [{
            "ProductGroup": "DOM",
            "ProductType": "CDA",
            "NumberOfShipments": extra['shipment_qty'],
            "PackageType":"BOX",
            "Payment":payment_type,
            "ShipmentWeight":{
                    "Unit": "KG",
                    "Value": 0.5
            },
            "ShipmentVolume": None,
            "NumberOfPieces": qty,
            "CashAmount": None,
            "ExtraCharges": None,
            "ShipmentDimensions": {
                "Length": 10,
                "Width": 10,
                "Height": 10,
                "Unit": "cm"
            },
            "Comments": "Handle with care."
        }
        ],
        "Status": "Ready"
    }

    # orders = logistic.orders.all().values_list('order_number', flat=True)
    # order_reference = ",".join(orders)
    aramex_transaction = {
        "Reference1": logistic_ref,
        "Reference2": ret,
        "Reference3": "",
        "Reference4": "",
        "Reference5": ""
    }
    data = {"ClientInfo": client_info, "LabelInfo": label_info,
            "Pickup": pickup, "Transaction": aramex_transaction}
    headers = {'Content-Type': 'application/json',
               'Accept': 'application/json'}
    # data = {'booking_items': booking_items, 'consignee': consignee,
    #         'consignor': consignor, 'parent_order': parent_order}
    try:
        resp = requests.post(url, data=json.dumps(data), headers=headers)
        json_response = resp.json()
        return json_response
    except Exception as e:
        print e
        return None

"""
Aramex API call for creating shipment.
"""
def aramex_shipment_create(logistic, seller, buyer, pickup_guid, payment_type, extra, third_party=None, ret=""):
    # pdb.set_trace()
    orders = logistic.orders.all().values_list('order_number', flat=True)
    order_reference = ",".join(orders) or 'ret-sha'

    client_info = {
        "UserName": settings.ARAMEX_USERNAME,
        "Password": settings.ARAMEX_PASSWORD,
        "Version": settings.ARAMEX_VERSION,
        "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
        "AccountPin": settings.ARAMEX_ACCOUNT_PIN,
        "AccountEntity": settings.ARAMEX_ACCOUNT_ENTITY,
        "AccountCountryCode": settings.ARAMEX_ACCOUNT_COUNTRY_CODE
    }

    label_info = {
        "ReportID": 9201,
        "ReportType": "URL"
    }
    shipment = {
        "Reference1": order_reference,
        "Reference2": "",
        "Reference3": "",
        "Shipper": {
            "Reference1": "",
            "Reference2": "",
            "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
            "PartyAddress": seller['address'],
            "Contact": seller['contact']
        },
        "Consignee": {
            "Reference1": "",
            "Reference2": "",
            "AccountNumber": "",
            "PartyAddress": buyer['address'],
            "Contact": buyer['contact']
        },

        "ShippingDateTime": "/Date(" + extra['shipping_time'] + "+0530)/",
        "DueDate": "/Date(" + extra['duedate_time'] + "+0530)/",

        "Comments": "Handle with care.",
        "PickupLocation": "",
        "Details": {
            "Dimensions": None,
            "ActualWeight": {
                "Unit": "KG",
                "Value": 0.5
            },
            "ChargeableWeight": None,
            "DescriptionOfGoods": "Books",
            "GoodsOriginCountry": "IN",
            "NumberOfPieces": 1,
            "ProductGroup": "DOM",
            "ProductType": "CDA",
            "PaymentType": payment_type,
            "PaymentOptions": "",
            "CustomsValueAmount": {"CurrencyCode": "INR", "Value": extra['total_amount']},
            "CashOnDeliveryAmount": {"CurrencyCode": "INR", "Value": extra['cod_amt']} if extra['cod'] else None,
            "Services": "CODS" if extra['cod'] else ""
        },
        "PickupGUID": pickup_guid
    }

    # "PickupGUID":"86aed70b-0e60-44b6-81f2-1260c2b80ea6"
    # }
    # ],

    aramex_transaction = {
        "Reference1": order_reference,
        "Reference2": ret,
        "Reference3": "",
        "Reference4": "",
        "Reference5": ""
    }

    if third_party:
        shipment['ThirdParty'] = third_party
    data = {"ClientInfo": client_info, "LabelInfo": label_info,
            "Shipments": [shipment], "Transaction": aramex_transaction}
    print json.dumps(data),' shipment data'
    headers = {'Content-Type': 'application/json',
               'Accept': 'application/json'}
    # data = {'booking_items': booking_items, 'consignee': consignee,
    #         'consignor': consignor, 'parent_order': parent_order}
    url = settings.ARAMEX_BASE_URL + settings.ARAMEX_CREATE_SHIPMENT_URL
    try:
        # pdb.set_trace()
        resp = requests.post(url, data=json.dumps(data), headers=headers)
        json_response = resp.json()
        return json_response
    except Exception as e:
        print e
        return None

"""
Aramex shipment label.
"""
@task
def save_aramex_pdf(source, dest):
    # pdb.set_trace()
    try:
        label_file = urllib.URLopener()
        label_file.retrieve(source, dest)
    except:
        save_aramex_pdf.apply_async(args=[source, dest], countdown=120)

"""
Delhivery Excel sheet creation.
We need to pass this sheet to delhivery for pickups.
"""
def create_delhivery_excel(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
        wb = Workbook()
        ws = wb.active
        # put headers for xcel
        ws['A1'] = 'Address of Pickup'
        ws['B1'] = 'City'
        ws['C1'] = 'State'
        ws['D1'] = 'Pincode'
        ws['E1'] = 'Date and Time of Pickup'
        ws['F1'] = 'Number of Packages'
        ws['G1'] = 'Contact No.'
        wb.save(directory + "/pickup.xlsx")
    elif not os.path.isfile(directory + "/pickup.xlsx"):
        wb = Workbook()
        ws = wb.active
        # put headers for xcel
        ws['A1'] = 'Address of Pickup'
        ws['B1'] = 'City'
        ws['C1'] = 'State'
        ws['D1'] = 'Pincode'
        ws['E1'] = 'Date and Time of Pickup'
        ws['F1'] = 'Number of Packages'
        ws['G1'] = 'Contact No.'
        wb.save(directory + "/pickup.xlsx")

"""
Updates the delhivery excel sheet.
"""
def update_delhivery_excel(pickup_excel_location, logistic_dict):
    wb = load_workbook(pickup_excel_location)
    ws = wb.active
    ini = ws.max_row + 1
    ws['A' + str(ini)] = logistic_dict['address']
    ws['B' + str(ini)] = logistic_dict['city']
    ws['C' + str(ini)] = logistic_dict['state']
    ws['D' + str(ini)] = logistic_dict['pincode']
    ws['E' + str(ini)] = timezone.now().date().strftime('%d-%m-%Y') + \
        ",  12:00 PM"
    ws['F' + str(ini)] = '1'
    ws['G' + str(ini)] = logistic_dict['phone']
    wb.save(pickup_excel_location)

"""
Returns the integer value defined in LOGISTICS_STATUS of logistics.models.
"""
def get_status_value(partner, status):
    from zap_apps.logistics.models import LOGISTICS_STATUS
    status_dict = {}
    status_list = LOGISTICS_STATUS.get(partner)
    status_dict = dict(status_list)
    return status_dict.get(status)

"""
Returns generic data needed for emails, etc.
"""
def get_products_template_data(orders):
    products = []
    for order in orders:
        products_email_dict = {}
        # product_titles += ords.product.title + ", "
        products_email_dict.update(
            {'productLink': settings.CURRENT_DOMAIN + '/#/product/' + str(order.product.id)})
        # pdb.set_trace()
        img = order.ordered_product.image.image.url
        products_email_dict.update(
            {'img': settings.CURRENT_DOMAIN + img})
        products_email_dict.update(
            {'productName': order.ordered_product.title})
        productSize = order.ordered_product.size
        products_email_dict.update({'productSize': productSize})
        products_email_dict.update(
            {'productColor': order.ordered_product.color})
        total_order_price = locale.format(
            "%.2f", order.final_price, grouping=True)
        productPrice = str(total_order_price)
        products_email_dict.update({'productQuantity': order.quantity})
        products_email_dict.update({'productPrice': productPrice})
        products.append(products_email_dict)
    return products

# 9am

"""
Works on top of logistics for orders.
API calls to corresponding logistics parnter.
Sends out necessary emails and saves packing slips, etc.
"""
# @periodic_task(run_every=(crontab(minute=30, hour=[2, 5])))
def pickup_orders_logistics(logistic_id=None):
    '''In this products with zapyle will never come
    '''

    # protocol = getattr(settings, "DEFAULT_HTTP_PROTOCOL", "http")
    zapemail = ZapEmail()
    if logistic_id:
        logistics_list = Logistics.objects.filter(id=logistic_id,
                                                  triggered_pickup=False, orders__isnull=False).distinct()
    else:
        logistics_list = Logistics.objects.filter(
            triggered_pickup=False, orders__isnull=False).distinct()

    if logistics_list:
        zapsms = ZapSms()
        todays_date = timezone.now().date().strftime('%d-%m-%Y')
        directory = settings.HOME_FOLDER + "/operations/" + todays_date

        after_two_days = timezone.now().date() + timezone.timedelta(2)
        after_two_days = after_two_days.strftime('%d-%m-%Y')
        future_directory = settings.HOME_FOLDER + \
            "/operations/" + after_two_days

        create_delhivery_excel(future_directory)

    delhivery_logistics_finder = {}
    sms_msg = "Heads up: We have scheduled a pickup from the address given in your listing. Please check your email for more information."

    if timezone.now().hour <= 2 and timezone.now().minute <= 30:
        aramex_shipment_time = timezone.now() + timezone.timedelta(hours=5)
        aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

    elif timezone.now().hour >= 11:
        aramex_shipment_time = datetime.combine(timezone.now().date() +
                                                timezone.timedelta(days=1), time(6, 5))
        aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

    else:
        aramex_shipment_time = timezone.now() + timezone.timedelta(hours=3, minutes=5)
        aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)



    aramex_extra = {
        'shipping_time': str(int((aramex_shipment_time.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
        'duedate_time': str(int((aramex_duedate_time.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
        'cod': False
    }

    for logistic in logistics_list:
        print logistic, '12345'

        consignee = logistic.consignee
        consignor = logistic.consignor

        orders = logistic.orders.all()
        products_desc = ""
        product_category = ""
        product_titles = ""
        total_amount = 0
        qty = 0
        for order in orders:
            # if order.cod:
            #     cod_amt += order.transaction.final_price
            products_desc = products_desc + " " + str(order.quantity) + " " + order.ordered_product.color + " " \
                + order.ordered_product.size + order.ordered_product.brand + \
                " " + order.ordered_product.product_category + ","

            total_amount += (order.final_price *
                             order.quantity)
            product_titles += order.ordered_product.title + ", "
            product_category += order.ordered_product.product_category + ", "
            qty += order.quantity
        products = get_products_template_data(orders)
        products_desc = products_desc.replace(products_desc[-1:], '.')
        product_titles = product_titles[
            :len(product_titles) - 2] + product_titles[-2:].replace(product_titles[-2:], ' ')
        product_category = product_category[
            :len(product_category) - 2] + product_category[-2:].replace(product_category[-2:], ' ')

        if logistic.consignor.user.user_type.name == "store_front":
            userTypeDecider = 1
        else:
            userTypeDecider = 0
        seller_email = ''
        if not consignor.user.user_type.name == 'zap_exclusive':
            seller_name = consignor.user.get_full_name()
            seller_email = consignor.user.email
            seller_phone = consignor.user.phone_number
        else:
            userdata = logistic.orders.all()[
                0].product.zapexclusiveuserdata_set.all()[0]
            seller_name = userdata.account_holder
            seller_email = userdata.email
            seller_phone = userdata.phone_number
        print logistic.pickup_partner,'code 2112'
        if logistic.pickup_partner == 'AR':
            #
            seller_add = {
                "Line1": consignor.address,
                "Line2": consignor.address2 or "",
                "Line3": "",
                "City": consignor.city,
                "StateOrProvinceCode": consignor.state.name,
                "PostCode": consignor.pincode,
                "CountryCode": "IN"
            }
            seller_contact = {
                "PersonName": consignor.name,
                "CompanyName": consignor.name,
                "PhoneNumber1": consignor.phone,
                "PhoneNumber2": "",
                "FaxNumber": "",
                "CellPhone": consignor.phone,
                "EmailAddress": consignor.user.email,
                "Type": ""
            }
            aramex_logistic_ref = str(logistic.id)
            # order_reference = ",".join(orders)

            # pdb.set_trace()
            if timezone.now().hour <= 6 and timezone.now().minute <= 30: # 0.00am to 12 pm
                aramex_pickup_datetime = datetime.combine(timezone.now().date(), time(12, 5))
                aramex_pickup_readytime = datetime.combine(timezone.now().date(), time(12, 10))
                aramex_pickup_lasttime = datetime.combine(timezone.now().date(), time(14))
                aramex_pickup_closetime = aramex_pickup_lasttime + timezone.timedelta(days=1)
            elif timezone.now().hour >= 11 and timezone.now().minute >= 30: # 5.00pm to 12am
                aramex_pickup_datetime = datetime.combine(timezone.now().date() +
                                                          timezone.timedelta(days=1), time(12))
                aramex_pickup_readytime = datetime.combine(timezone.now().date() +
                                                           timezone.timedelta(days=1), time(12, 5))
                aramex_pickup_lasttime = aramex_pickup_readytime + \
                    timezone.timedelta(hours=7, minutes=30)
                aramex_pickup_closetime = aramex_pickup_lasttime + \
                    timezone.timedelta(days=1)
            else:
                aramex_pickup_datetime = timezone.now() + timezone.timedelta(hours=5, minutes=35)
                aramex_pickup_readytime = timezone.now() + timezone.timedelta(hours=5, minutes=40)
                aramex_pickup_lasttime = timezone.now() + timezone.timedelta(hours=8, minutes=5)
                aramex_pickup_closetime = aramex_pickup_lasttime + timezone.timedelta(days=1)
            
            print aramex_pickup_datetime,'aramex_pickup_datetime'
            print aramex_pickup_readytime,'aramex_pickup_readytime'
            print aramex_pickup_lasttime,'aramex_pickup_lasttime'
            print aramex_pickup_closetime,'aramex_pickup_closetime'
            print timezone.now().hour,'timezone.now().hour'
            aramex_pickup_extra = {
                'aramex_pickup_datetime': str(int((aramex_pickup_datetime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
                'aramex_pickup_readytime': str(int((aramex_pickup_readytime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
                'aramex_pickup_lasttime': str(int((aramex_pickup_lasttime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
                'aramex_pickup_closetime': str(int((aramex_pickup_closetime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),

                'shipment_qty': 1
            }
            # pdb.set_trace()
            json_response = aramex_pickup_create(
                aramex_logistic_ref, qty, seller_add, seller_contact, "3", aramex_pickup_extra)
            print json_response,'code 5476'
            if json_response and not json_response['HasErrors']:
                aramex_pickup_guid = json_response['ProcessedPickup']['GUID']
                seller_shipment = {
                    'address': seller_add,
                    'contact': seller_contact
                }
                zapyle_shipment = {
                    'address': {
                        "Line1": "Zapyle",
                        "Line2": "CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli",
                        "Line3": "",
                        "City": "Bangalore",
                        "StateOrProvinceCode": "Karnataka",
                        "PostCode": "560016",
                        "CountryCode": "IN"
                    },
                    'contact': {
                        "PersonName": "Zapyle",
                        "CompanyName": "Zapyle",
                        "PhoneNumber1": "07337880282",
                        "PhoneNumber2": "",
                        "FaxNumber": "",
                        "CellPhone": "9663596594",
                        "EmailAddress": "hello@zapyle.com",
                        "Type": ""
                    }
                }
                third_party = {
                    "Reference1": "",
                    "Reference2": "",
                    "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
                    "PartyAddress": zapyle_shipment['address'],
                    "Contact": zapyle_shipment['contact']
                }
                aramex_extra['total_amount'] = total_amount

                shipment_resp = aramex_shipment_create(
                    logistic, seller_shipment, zapyle_shipment, aramex_pickup_guid, "3", aramex_extra, third_party)
                # pdb.set_trace()
                print shipment_resp,'shipment_response'
                if shipment_resp and not shipment_resp['HasErrors']:
                    status = 0
                    shipment_data = shipment_resp['Shipments'][0]
                    pickup_data = {'logistics': logistic.id, 'waybill': shipment_data['ID'],
                                   'log_status': status, 'track': True, 'whole_response': json.dumps(shipment_resp),
                                   'logistics_ref': str(logistic.id), 'partner': 'AR', 'extra': json.dumps(json_response['ProcessedPickup'])}
                    print 'success11'
                    pickup_log_serializer = LogisticsLogSerializer(
                        data=pickup_data)
                    if pickup_log_serializer.is_valid():
                        pickup_log_serializer.save()
                        logistic.triggered_pickup = True
                        logistic.save()
                    else:
                        print pickup_log_serializer.errors,'23433'
                        # aramex_pdf_url = settings.HOME_FOLDER + "/operations/" + \
                        #     after_two_days + "/packing_slip_seller_" + \
                        #     shipment_data['ID'] + ".pdf"
                        # save_aramex_pdf.delay(shipment_data['ShipmentLabel']['LabelURL'], aramex_pdf_url)
                    try:
                        template_data = {'userTypeDecider': userTypeDecider, 'seller_name': seller_name,
                                         'product_titles': product_titles, 'buyer_name': logistic.consignee.user.get_full_name(), 'products': products}
                        html = settings.SELLER_GENERIC
                        html_body = render_to_string(
                            html['html'], template_data)

                        zapemail.send_email_alternative(html[
                            'subject'], settings.FROM_EMAIL, seller_email, html_body)
                    except:
                        print "email not sent"
                    try:
                        zapsms.send_sms(seller_phone, sms_msg)
                    except Exception as e:
                        print e
            else:
                # send sms to shafi when getting error from aramex #
                r = requests.get("http://bhashsms.com/api/sendmsg.php?user=Zapyle&pass=zapyle@123&sender=ZAPYLE&phone=9895685141&text={}&priority=ndnd&stype=normal".format(json_response['Notifications'][0]['Message']))
        elif logistic.pickup_partner == 'DL':
            # pdb.set_trace()

            waybill_json = delhivery_create_waybill()
            seller_add = {'add': consignor.name + ', ' + consignor.address, 'city': consignor.city,
                          'country': consignor.country, 'name': settings.DELHIVERY_PICKUP_NAME, 'phone': consignor.phone, 'pin': consignor.pincode}

            packaging_shipments = {'waybill': waybill_json, 'name': 'Zapyle', 'order': logistic.id, 'products_desc': products_desc,
                                   'order_date': timezone.now().isoformat(), 'payment_mode': 'Pre-paid', 'total_amount': total_amount,
                                   'add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore', 'state': 'Karnataka', 'pin': '560016',
                                   'phone': '07337880282', 'weight': "",  'product_quantity': 1}

            pickup_json_response, content = delhivery_create(
                packaging_shipments, seller_add, waybill_json)

            if (pickup_json_response) and ('success' in pickup_json_response) and (pickup_json_response['success'] is True):
                status = get_status_value(
                    'DL', pickup_json_response["packages"][0]["status"])
                pickup_response_data = {'logistics': logistic.id, 'waybill': waybill_json, 'log_status': status, 'track': True,
                                        'whole_response': content, 'logistics_ref': pickup_json_response["upload_wbn"], 'partner': 'DL'}
                print 'success11'
                pickup_log_serializer = LogisticsLogSerializer(
                    data=pickup_response_data)
                if pickup_log_serializer.is_valid():
                    pickup_log_serializer.save()
                    logistic.triggered_pickup = True
                    logistic.save()

                logistic_dict = {'address': logistic.consignor.name + ", " + logistic.consignor.address, 'city': logistic.consignor.city,
                                 'state': logistic.consignor.state.name, 'pincode': logistic.consignor.pincode, 'phone': logistic.consignor.phone}

                update_delhivery_excel(
                    future_directory+"/pickup.xlsx", logistic_dict)
                try:
                    zapsms.send_sms(seller_phone, sms_msg)
                except Exception as e:
                    print e

                delhivery_logistics_finder.update({waybill_json: logistic})

                # SAVE HTML FOR PDF > SAVE IT IN EXCEL > SEND IT TO SELLER

            else:
                if 'error' in pickup_json_response:
                    error_message = pickup_json_response['error']
                else:
                    error_message = pickup_json_response["rmk"]
                pickup_response_data = {'partner': 'DL', 'logistics_ref': pickup_json_response["upload_wbn"], 'error_message': error_message}
                pickup_log_serializer = LogisticsLogSerializer(
                    data=pickup_response_data)
                if pickup_log_serializer.is_valid():
                    pickup_log_serializer.save()

            # DELHIVERY PICK UP ENDS

        elif logistic.pickup_partner == 'ZP':
            # ######pdb.set_trace()
            pickup_data = {'logistics': logistic.id, 'waybill': 'ZAPPICK' + str(
                logistic.id), 'log_status': 0, 'track': True, 'whole_response': 'Schedule a pickup and change to manifested once Amzad is informed.',
                'logistics_ref': str(logistic.id), 'partner': 'ZP'}
            print 'success11'
            pickup_log_serializer = LogisticsLogSerializer(
                data=pickup_data)
            if pickup_log_serializer.is_valid():
                pickup_log_serializer.save()
                logistic.triggered_pickup = True
                logistic.save()

            try:
                html = settings.SELLER_GENERIC
                template_data = {'userTypeDecider': userTypeDecider, 'seller_name': seller_name,
                                 'product_titles': product_titles, 'buyer_name': logistic.consignee.user.get_full_name(), 'products': products}
                html_body = render_to_string(
                    html['html'], template_data)

                zapemail.send_email_alternative(html[
                    'subject'], settings.FROM_EMAIL, seller_email, html_body)
            except Exception as e:
                zaplogging.delay("Zapyle-order-create-error-{}".format(
                    str(e)) if settings.CELERY_USE else zaplogging("Zapyle-order-create-error-{}".format(str(e))))
            try:
                zapsms.send_sms(seller_phone, sms_msg)
            except Exception as e:
                print e

            # ZAPYLE PICK UP ENDS

            # SAVE DIFF HTML FOR PDF > SEND IT TO SELLER

        # PICKUP FOR PARCELLED
        elif logistic.pickup_partner == 'PR':
            # ######pdb.set_trace()

            # file.write("+++______response________+++")
            # file.write(str(json_response['data']))
            # file.write(str(json_response['meta']['code']))

            pr_logistic_id = logistic.id

            pr_booking_item = {}
            pr_booking_item.update({'category': product_category})
            pr_booking_item.update({'qty': qty})
            pr_booking_item = [pr_booking_item]
            pr_consignee = {'address': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore',
                            'country': 'India', 'name': 'Zapyle', 'phone': '07337880282', 'pincode': '560016', 'state': 'Karnataka'}
            pr_consignor = {'address': consignor.address, 'city': consignor.city, 'country': consignor.country, 'name': consignor.name,
                            'phone': consignor.phone, 'pincode': consignor.pincode, 'state': consignor.state.name, 'email': seller_email}

            json_response = parcelled_create(
                str(pr_logistic_id), pr_booking_item, pr_consignee, pr_consignor)

            # ######pdb.set_trace()
            if (json_response['meta']['code'] == 200 or json_response['meta']['code'] == 201):
                status = get_status_value(
                    'PR', json_response['data']['bookings'][0]['status'])
                pickup_data = {'logistics': logistic.id, 'waybill': json_response['data']['bookings'][0]['booking_id'],
                               'log_status': status, 'track': True, 'whole_response': json.dumps(json_response),
                               'logistics_ref': json_response['data']['bookings'][0]['parent_order']['order_id'], 'partner': 'PR'}
                print 'success11'
                pickup_log_serializer = LogisticsLogSerializer(
                    data=pickup_data)
                if pickup_log_serializer.is_valid():
                    pickup_log_serializer.save()
                    logistic.triggered_pickup = True
                    logistic.save()
                try:
                    template_data = {'userTypeDecider': userTypeDecider, 'seller_name': seller_name,
                                     'product_titles': product_titles, 'buyer_name': logistic.consignee.user.get_full_name(), 'products': products}
                    html = settings.SELLER_GENERIC
                    html_body = render_to_string(
                        html['html'], template_data)

                    zapemail.send_email_alternative(html[
                        'subject'], settings.FROM_EMAIL, seller_email, html_body)
                except:
                    print "email not sent"
                try:
                    zapsms.send_sms(seller_phone, sms_msg)
                except Exception as e:
                    print e
            # LOG ERRORS

        # PARCELLED PICK UP ENDS

    # CALL DELHIVERY PACKING SLIP API AND SEND EMAILS WITH DELHIVERY BILL
    # ATTACHED
    waybill_keys = delhivery_logistics_finder.keys()
    waybill_str = ",".join(waybill_keys)
    # ######pdb.set_trace()
    packing_slip = delhivery_packing_slip(waybill_str)
    if packing_slip and 'packages' in packing_slip:
        for packages in packing_slip['packages']:
            logistics_data = delhivery_logistics_finder.get(packages['wbn'])
            if logistics_data.logistics.consignor.user.user_type.name == "store_front":
                userTypeDecider = 1
            else:
                userTypeDecider = 0

            order_number = ""
            for single_order in logistics_data.orders.all():
                order_number += single_order.order_number + ", "
            order_number = order_number[:-2]
            coc_code = 'X9'
            oid_barcode = decode_base64(packages['oid_barcode'])
            wbn_barcode = decode_base64(packages['barcode'])
            oid_filename = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/DL_pickup_oid_bar_" + \
                packages['wbn'] + ".png"
            wbn_filename = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/DL_pickup_wbn_bar_" + \
                packages['wbn'] + ".png"
            delhivery_logo_full = packages['delhivery_logo']
            delhivery_logo_raw = delhivery_logo_full.split(":")[1]
            delhivery_logo = "https://track.delhivery.com" + delhivery_logo_raw
            with open(oid_filename, 'wb') as f:
                f.write(oid_barcode)
            f.close()
            with open(wbn_filename, 'wb') as f:
                f.write(wbn_barcode)
            f.close()

            waybill_barcode = settings.CURRENT_DOMAIN + "/logmedia/" + \
                after_two_days + "/DL_pickup_wbn_bar_" + \
                packages['wbn'] + ".png"
            order_barcode = settings.CURRENT_DOMAIN + "/logmedia/" + \
                after_two_days + "/DL_pickup_oid_bar_" + \
                packages['wbn'] + ".png"

            packing_slip_template_data = {'delhivery_logo': delhivery_logo, 'waybill_barcode': waybill_barcode, 'order_barcode': order_barcode,
                                          'order_number': order_number, 'coc_code': coc_code, 'from_name': logistics_data.consignor.name, 'from_add': logistics_data.consignor.address,
                                          'from_city': logistics_data.consignor.city, 'from_state': logistics_data.consignor.state.name, 'from_pincode': logistics_data.consignor.pincode,
                                          'from_phone': logistics_data.consignor.phone, 'to_name': 'Zapyle', 'to_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli',
                                          'to_city': 'Bangalore', 'to_state': 'Karnataka', 'to_pincode': '560016', 'to_phone': '07337880282', 'package_product': packages['prd'],
                                          'package_price': str(packages['rs']), 'userTypeDecider': userTypeDecider, 'cod_decider': 0}
            packing_slip_html_body = render_to_string(
                "delhivery_packing_slip.html", packing_slip_template_data)

            Html_file = open(settings.HOME_FOLDER + "/operations/" +
                             after_two_days + "/DL_pickup_" + packages['wbn'] + '.html', "w")
            Html_file.write(packing_slip_html_body)
            Html_file.close()

            html_url = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/DL_pickup_" + packages['wbn'] + ".html"
            pdf_url = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/packing_slip_seller_" + \
                packages['wbn'] + ".pdf"
            config = pdfkit.configuration(
                wkhtmltopdf=settings.WKHTMLTOPDF_PATH)
            pdfkit.from_string(packing_slip_html_body,
                               pdf_url, configuration=config)
            template_data = {'userTypeDecider': userTypeDecider, 'seller_name': logistics_data.consignor.user.get_full_name(
            ), 'product_titles': product_titles, 'buyer_name': logistics_data.consignee.user.get_full_name(), 'products': products}

            html = settings.SELLER_NOTIF
            html_body = render_to_string(
                html['html'], template_data)

            attachment = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/packing_slip_seller_" + \
                packages['wbn'] + ".pdf"
            attachment_name = "packing_slip_" + packages['wbn'] + ".pdf"

            zapemail.send_email_alternative(html[
                'subject'], settings.FROM_EMAIL, logistics_data.consignor.user.email, html_body, attachment, attachment_name)
    # DELHIVERY PACKING SLIP ENDS
        # ALL PICKUP ENDS


# def zapexclusive_inhouse_delivery():


#@periodic_task(run_every=(crontab(minute=31, hour=2)))
"""
Works on top of logistics for returns.
API calls to corresponding logistics parnter.
Sends out necessary emails and saves packing slips, etc.
"""
# @periodic_task(run_every=(crontab(minute=31, hour=[2, 5])))
def pickup_returns_logistics(id=None):
    # pdb.set_trace()
    zapemail = ZapEmail()
    Q1 = Q(triggered_pickup=False)
    Q2 = Q(returns__isnull=False)
    # Q3 = Q(product_delivery_partner='DL')
    if id:
        return_logistics = Logistics.objects.filter(
            Q1 & Q2 & Q(id=id)).distinct()
    else:
        return_logistics = Logistics.objects.filter(Q1 & Q2).distinct()
    # url = settings.DELHIVERY_BASE_URL + \
    #     '/cmu/push/json/?token=' + settings.DELHIVERY_API_KEY
    # print return_logistics, 'return_logistics'

    if return_logistics:
        zapsms = ZapSms()

        todays_date = timezone.now().date().strftime('%d-%m-%Y')
        directory = settings.HOME_FOLDER + "/operations/" + todays_date

        after_two_days = timezone.now().date() + timezone.timedelta(2)
        after_two_days = after_two_days.strftime('%d-%m-%Y')
        future_directory = settings.HOME_FOLDER + \
            "/operations/" + after_two_days

        create_delhivery_excel(future_directory)

    delhivery_logistics_finder = {}

    other_return_sms_msg = "Hey Zapyler! Your request for the return of your Zapyle purchase has been approved. Please hand it over for pick up to our Zap rep and check your mail for the further details."

    if timezone.now().hour <= 2 and timezone.now().minute <= 30:
        aramex_shipment_time = timezone.now() + timezone.timedelta(hours=5)
        aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

    elif timezone.now().hour >= 11:
        aramex_shipment_time = datetime.combine(timezone.now().date() +
                                                timezone.timedelta(days=1), time(6, 5))
        aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

    else:
        aramex_shipment_time = timezone.now() + timezone.timedelta(hours=3, minutes=5)
        aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

    aramex_extra = {
        'shipping_time': str(int((aramex_shipment_time.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
        'duedate_time': str(int((aramex_duedate_time.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
        'cod': False
    }

    for logistic in return_logistics:
        # print logistics, 'logistics111'
        consignee = logistic.consignee
        consignor = logistic.consignor

        returns = logistic.returns.all()
        # print returns, ' logistics111 returns'
        products_desc = ""
        product_category = ""
        total_amount = 0
        qty = 0
        for ret in returns:
            products_desc = products_desc + " " + \
                str(ret.order_id.quantity) + " " + \
                ret.order_id.ordered_product.product_category + ","
            total_amount += (ret.order_id.final_price *
                             ret.order_id.quantity)
            # total_quantity += ret.order_id.quantity
            product_category += ret.order_id.ordered_product.product_category + \
                ", "
            qty += ret.order_id.quantity
        product_category = product_category[
            :len(product_category) - 2] + product_category[-2:].replace(product_category[-2:], ' ')
        products_desc = products_desc.replace(products_desc[-1:], '.')

        # if not logistic.pickup_partner:
        if logistic.pickup_partner == 'AR':
            #
            seller_add = {
                "Line1": consignor.address,
                "Line2": consignor.address2 or "",
                "Line3": "",
                "City": consignor.city,
                "StateOrProvinceCode": consignor.state.name,
                "PostCode": consignor.pincode,
                "CountryCode": "IN"
            }
            seller_contact = {
                "PersonName": consignor.name,
                "CompanyName": consignor.name,
                "PhoneNumber1": consignor.phone,
                "PhoneNumber2": "",
                "FaxNumber": "",
                "CellPhone": consignor.phone,
                "EmailAddress": consignor.user.email,
                "Type": ""
            }
            aramex_logistic_ref = str(logistic.id)
            aramex_ret_ref = "Return - " + \
                str(logistic.returns.all().values_list('id', flat=True))

            # pdb.set_trace()
            if timezone.now().hour <= 6 and timezone.now().minute <= 30: # 0.00am to 12 pm
                aramex_pickup_datetime = datetime.combine(timezone.now().date(), time(12, 5))
                aramex_pickup_readytime = datetime.combine(timezone.now().date(), time(12, 10))
                aramex_pickup_lasttime = datetime.combine(timezone.now().date(), time(14))
                aramex_pickup_closetime = aramex_pickup_lasttime + timezone.timedelta(days=1)
            elif timezone.now().hour >= 11 and timezone.now().minute >= 30: # 5.00pm to 12am
                aramex_pickup_datetime = datetime.combine(timezone.now().date() +
                                                          timezone.timedelta(days=1), time(12))
                aramex_pickup_readytime = datetime.combine(timezone.now().date() +
                                                           timezone.timedelta(days=1), time(12, 5))
                aramex_pickup_lasttime = aramex_pickup_readytime + \
                    timezone.timedelta(hours=7, minutes=30)
                aramex_pickup_closetime = aramex_pickup_lasttime + \
                    timezone.timedelta(days=1)
            else:
                aramex_pickup_datetime = timezone.now() + timezone.timedelta(hours=5, minutes=35)
                aramex_pickup_readytime = timezone.now() + timezone.timedelta(hours=5, minutes=40)
                aramex_pickup_lasttime = timezone.now() + timezone.timedelta(hours=8, minutes=5)
                aramex_pickup_closetime = aramex_pickup_lasttime + timezone.timedelta(days=1)

            aramex_pickup_extra = {
                'aramex_pickup_datetime': str(int((aramex_pickup_datetime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
                'aramex_pickup_readytime': str(int((aramex_pickup_readytime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
                'aramex_pickup_lasttime': str(int((aramex_pickup_lasttime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
                'aramex_pickup_closetime': str(int((aramex_pickup_closetime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),

                'shipment_qty': 1
            }

            json_response = aramex_pickup_create(
                aramex_logistic_ref, qty, seller_add, seller_contact, "3", aramex_pickup_extra, aramex_ret_ref)
            if json_response and not json_response['HasErrors']:
                aramex_pickup_guid = json_response['ProcessedPickup']['GUID']
                seller_shipment = {
                    'address': seller_add,
                    'contact': seller_contact
                }
                zapyle_shipment = {
                    'address': {
                        "Line1": "Zapyle",
                        "Line2": "CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli",
                        "Line3": "",
                        "City": "Bangalore",
                        "StateOrProvinceCode": "Karnataka",
                        "PostCode": "560016",
                        "CountryCode": "IN"
                    },
                    'contact': {
                        "PersonName": "Zapyle",
                        "CompanyName": "Zapyle",
                        "PhoneNumber1": "07337880282",
                        "PhoneNumber2": "",
                        "FaxNumber": "",
                        "CellPhone": "9663596594",
                        "EmailAddress": "hello@zapyle.com",
                        "Type": ""
                    }
                }
                third_party = {
                    "Reference1": "",
                    "Reference2": "",
                    "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
                    "PartyAddress": zapyle_shipment['address'],
                    "Contact": zapyle_shipment['contact']
                }
                aramex_extra['total_amount'] = total_amount
                # pdb.set_trace()
                shipment_resp = aramex_shipment_create(
                    logistic, seller_shipment, zapyle_shipment, aramex_pickup_guid, "3", aramex_extra, third_party)
                if shipment_resp and not shipment_resp['HasErrors']:
                    status = 0
                    shipment_data = shipment_resp['Shipments'][0]
                    pickup_data = {'logistics': logistic.id, 'waybill': shipment_data['ID'], 'returns': True,
                                   'log_status': status, 'track': True, 'whole_response': json.dumps(shipment_resp),
                                   'logistics_ref': str(logistic.id), 'partner': 'AR', 'extra': json.dumps(json_response['ProcessedPickup'])}
                    print 'success11'
                    pickup_log_serializer = LogisticsLogSerializer(
                        data=pickup_data)
                    if pickup_log_serializer.is_valid():
                        pickup_log_serializer.save()
                        logistic.triggered_pickup = True
                        logistic.save()
                        logistic.returns.update(return_status='pickup_in_process')
                        # aramex_pdf_url = settings.HOME_FOLDER + "/operations/" + \
                        #     after_two_days + "/packing_slip_seller_" + \
                        #     shipment_data['ID'] + ".pdf"
                        # save_aramex_pdf.delay(shipment_data['ShipmentLabel']['LabelURL'], aramex_pdf_url)
                    try:
                        return_template_data = {
                            'buyer_name': logistic.consignor.user.get_full_name()}

                        html = settings.RETURN_AFTER_SCHEDULE
                        html_body = render_to_string(
                            html['html'], return_template_data)

                        zapemail.send_email_alternative(html[
                            'subject'], settings.FROM_EMAIL, logistic.consignor.user.email, html_body)
                    except:
                        print "return doc not sent"
                    try:

                        zapsms.send_sms(
                            logistic.consignor.user.phone_number, other_return_sms_msg)
                    except Exception as e:
                        print e

        elif logistic.pickup_partner == 'DL':
            waybill_json = delhivery_create_waybill()

            pickup_location = {'add': consignor.address, 'city': consignor.city,
                               'name': settings.DELHIVERY_PICKUP_NAME, 'phone': consignor.phone, 'pin': consignor.pincode}
            shipments = {'waybill': waybill_json, 'name': consignee.name, 'order': logistic.id, 'products_desc': products_desc,
                         'order_date': timezone.now().isoformat(), 'payment_mode': 'Pre-paid', 'total_amount': total_amount,
                         'add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore', 'state': 'Karnataka', 'pin': '560016',
                         'phone': '07337880282', 'weight': "",  'product_quantity': 1, 'return_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli',
                         'return_city': 'Bangalore', 'return_name': 'Zapyle', 'return_phone': '07337880282', 'return_pin': '560016', 'return_state': 'Karnataka'}

            pickup_json_response, content = delhivery_create(
                shipments, pickup_location, waybill_json)

            if (pickup_json_response) and ('success' in pickup_json_response) and (pickup_json_response['success'] is True):
                status = get_status_value(
                    'DL', pickup_json_response["packages"][0]["status"])
                pickup_response_data = {'logistics': logistic.id, 'waybill': waybill_json, 'log_status': status, 'track': True,
                                        'whole_response': content, 'logistics_ref': pickup_json_response["upload_wbn"], 'partner': 'DL', 'returns': True}
                print 'success11'
                pickup_log_serializer = LogisticsLogSerializer(
                    data=pickup_response_data)
                if pickup_log_serializer.is_valid():
                    pickup_log_serializer.save()
                    logistic.triggered_pickup = True
                    logistic.save()
                    logistic.returns.update(return_status='pickup_in_process')
                logistic_dict = {'address': logistic.consignor.name + ", " + logistic.consignor.address, 'city': logistic.consignor.city,
                                 'state': logistic.consignor.state.name, 'pincode': logistic.consignor.pincode, 'phone': logistic.consignor.phone}

                update_delhivery_excel(
                    future_directory+"/pickup.xlsx", logistic_dict)
                try:
                    DL_sms_msg = "Hey Zapyler! Your request for the return of your Zapyle purchase has been approved. Please hand it over for pick up to our Zap rep and check your mail for the returning document."
                    zapsms.send_sms(
                        logistic.consignor.user.phone_number, DL_sms_msg)
                except Exception as e:
                    print e

                delhivery_logistics_finder.update({waybill_json: logistic})

                # SAVE HTML FOR PDF > SAVE IT IN EXCEL > SEND IT TO SELLER

            else:
                if 'error' in pickup_json_response:
                    error_message = pickup_json_response['error']
                else:
                    error_message = pickup_json_response["rmk"]
                pickup_response_data = {'partner': 'DL', 'logistics_ref': pickup_json_response[
                    "upload_wbn"], 'error_message': error_message}
                pickup_log_serializer = LogisticsLogSerializer(
                    data=pickup_response_data)
                if pickup_log_serializer.is_valid():
                    pickup_log_serializer.save()

        elif logistic.pickup_partner == 'ZP':

            pickup_data = {'logistics': logistic.id, 'waybill': 'ZAPRETPICK' + str(logistic.id), 'track': True,
                           'log_status': 0, 'track': True, 'whole_response': 'Schedule a pickup and change to manifested once Amzad is informed.',
                           'logistics_ref': str(logistic.id), 'partner': 'ZP', 'returns': True}
            print 'success11'
            pickup_log_serializer = LogisticsLogSerializer(
                data=pickup_data)
            if pickup_log_serializer.is_valid():
                pickup_log_serializer.save()
                logistic.triggered_pickup = True
                logistic.save()
                logistic.returns.update(return_status='pickup_in_process')
                try:
                    return_template_data = {
                        'buyer_name': logistic.consignor.user.get_full_name()}

                    # return_html_body = render_to_string(
                    #     "other_return.html", return_template_data)
                    html = settings.RETURN_AFTER_SCHEDULE
                    html_body = render_to_string(
                        html['html'], return_template_data)

                    zapemail.send_email_alternative(html[
                        'subject'], settings.FROM_EMAIL, logistic.consignor.user.email, html_body)

                    # email = EmailMultiAlternatives(subject="Follow Up on Returns - Zapyle",
                    #                              from_email="hello@zapyle.com", to=[logistic.consignor.user.email], body=return_html_body)
                    # email.attach_alternative(return_html_body, "text/html")
                    # email.send()

                except:
                    print "return doc not sent"
                try:

                    zapsms.send_sms(
                        logistic.consignor.user.phone_number, other_return_sms_msg)
                except Exception as e:
                    print e

        elif logistic.pickup_partner == 'PR':
            pr_logistic_id = logistic.id

            pr_booking_item = {}
            pr_booking_item.update({'category': product_category})
            pr_booking_item.update({'qty': qty})

            pr_booking_item = [pr_booking_item]
            pr_consignee = {'address': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore',
                            'country': 'India', 'name': 'Zapyle', 'phone': '07337880282', 'pincode': '560016', 'state': 'Karnataka'}
            pr_consignor = {'address': consignor.address, 'city': consignor.city, 'country': consignor.country, 'name': consignor.name,
                            'phone': consignor.phone, 'pincode': consignor.pincode, 'state': consignor.state.name, 'email': consignor.user.email}

            json_response = parcelled_create(
                str(pr_logistic_id), pr_booking_item, pr_consignee, pr_consignor)

            if (json_response['meta']['code'] == 200 or json_response['meta']['code'] == 201):
                status = get_status_value(
                    'PR', json_response['data']['bookings'][0]['status'])
                pickup_data = {'logistics': logistic.id, 'waybill': json_response['data']['bookings'][0]['booking_id'],
                               'log_status': status, 'track': True, 'whole_response': json.dumps(json_response),
                               'logistics_ref': json_response['data']['bookings'][0]['parent_order']['order_id'], 'partner': 'PR', 'returns': True}
                print 'success11'
                pickup_log_serializer = LogisticsLogSerializer(
                    data=pickup_data)
                if pickup_log_serializer.is_valid():
                    pickup_log_serializer.save()
                    logistic.returns.update(return_status='pickup_in_process')
                    logistic_update_serializer = LogisticsSerializer(
                        data={'triggered_pickup': True})
                    if logistic_update_serializer.is_valid():
                        logistic_update_serializer.update(
                            logistic, logistic_update_serializer.validated_data)
                try:
                    return_template_data = {
                        'buyer_name': logistic.consignor.user.get_full_name()}

                    html = settings.RETURN_AFTER_SCHEDULE
                    html_body = render_to_string(
                        html['html'], return_template_data)

                    zapemail.send_email_alternative(html[
                        'subject'], settings.FROM_EMAIL, logistic.consignor.user.email, html_body)
                except:
                    print "return doc not sent"
                try:

                    zapsms.send_sms(
                        logistic.consignor.user.phone_number, other_return_sms_msg)
                except Exception as e:
                    print e

            # LOG ERROR

    waybill_keys = delhivery_logistics_finder.keys()
    waybill_str = ",".join(waybill_keys)
    # ######pdb.set_trace()
    packing_slip = delhivery_packing_slip(waybill_str)
    if packing_slip and 'packages' in packing_slip:
        for packages in packing_slip['packages']:

            logistics_data = delhivery_logistics_finder.get(packages['wbn'])

            order_number = ""
            for returns in logistics_data.returns.all():
                order_number += returns.order_id.order_number + ", "
            order_number = order_number[:-2]
            coc_code = 'X9'

            oid_barc = packages['oid_barcode']
            wbn_barc = packages['barcode']
            oid_barcode = decode_base64(oid_barc)
            wbn_barcode = decode_base64(wbn_barc)
            oid_filename = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/DL_ret_pickup_oid_bar_" + \
                packages['wbn'] + ".png"
            wbn_filename = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/DL_ret_pickup_wbn_bar_" + \
                packages['wbn'] + ".png"
            delhivery_logo_full = packages['delhivery_logo']
            delhivery_logo_raw = delhivery_logo_full.split(":")[1]
            delhivery_logo = "https://track.delhivery.com" + delhivery_logo_raw
            with open(oid_filename, 'wb') as f:
                f.write(oid_barcode)
            f.close()
            with open(wbn_filename, 'wb') as f:
                f.write(wbn_barcode)
            f.close()

            # ########pdb.set_trace()

            waybill_barcode = settings.CURRENT_DOMAIN + "/logmedia/" + \
                after_two_days + "/DL_ret_pickup_wbn_bar_" + \
                packages['wbn'] + ".png"
            order_barcode = settings.CURRENT_DOMAIN + "/logmedia/" + \
                after_two_days + "/DL_ret_pickup_oid_bar_" + \
                packages['wbn'] + ".png"
            packing_slip_template_data = {'delhivery_logo': delhivery_logo, 'waybill_barcode': waybill_barcode, 'order_barcode': order_barcode,
                                          'order_number': order_number, 'coc_code': coc_code, 'from_name': logistics_data.consignor.name, 'from_add': logistics_data.consignor.address,
                                          'from_city': logistics_data.consignor.city, 'from_state': logistics_data.consignor.state.name, 'from_pincode': logistics_data.consignor.pincode,
                                          'from_phone': logistics_data.consignor.phone, 'to_name': 'Zapyle', 'to_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'to_city': 'Bangalore',
                                          'to_state': 'Karnataka', 'to_pincode': '560016', 'to_phone': '07337880282', 'package_product': packages['prd'],
                                          'package_price': str(packages['rs'])}
            packing_slip_html_body = render_to_string(
                "delhivery_return_doc.html", packing_slip_template_data)

            Html_file = open(settings.HOME_FOLDER + "/operations/" +
                             after_two_days + "/DL_ret_pickup_" + packages['wbn'] + '.html', "w")
            Html_file.write(packing_slip_html_body)
            Html_file.close()

            html_url = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/DL_ret_pickup_" + packages['wbn'] + ".html"
            pdf_url = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/ret_doc_" + packages['wbn'] + ".pdf"
            config = pdfkit.configuration(
                wkhtmltopdf=settings.WKHTMLTOPDF_PATH)
            pdfkit.from_string(packing_slip_html_body,
                               pdf_url, configuration=config)

            return_template_data = {
                'buyer_name': logistics_data.consignor.user.get_full_name()}

            html = settings.DELHIVERY_RETURN_DOC
            html_body = render_to_string(
                html['html'], return_template_data)
            attachment = settings.HOME_FOLDER + "/operations/" + \
                after_two_days + "/ret_doc_" + packages['wbn'] + ".pdf"
            attachment_name = "returns_" + packages['wbn'] + ".pdf"
            zapemail.send_email_alternative(html[
                'subject'], settings.FROM_EMAIL, logistics_data.consignor.user.email, html_body, attachment, attachment_name)

"""
Updates order status according to logistic log status.
"""
def update_order_status(log_status, logistic_log):
    # pdb.set_trace()
    order_status = ''
    if not logistic_log.returns:
        if logistic_log.pickup:
            if log_status == 2:
                order_status = 'picked_up'
            elif log_status == 4:
                order_status = 'verification'
        else:
            if logistic_log.rejected:
                if log_status == 2:
                    order_status = 'return_shipped'
                elif log_status == 4:
                    order_status = 'returned'
                    logistic_log.logistics.orders.update(product_verification='rejected_shipped')
                elif log_status == 5:
                    order_status = 'on_hold'
            else:
                if log_status == 2:
                    order_status = 'on_the_way_to_you'
                elif log_status == 4:
                    order_status = 'delivered'
                elif log_status == 5:
                    order_status = 'on_hold'
        logistic_log.log_status = log_status
        logistic_log.save()

    else:
        if logistic_log.pickup:
            if log_status == 2:
                order_status = 'picked_up'
            elif log_status == 4:
                order_status = 'verification'
        # if (logistic_log.pickup and logistic_log.logistics.returns.all()[0].order_id.ordered_product.with_zapyle) or (not logistic_log.pickup):
        #     if log_status == 4:
        #         order_status = 'returned'
    if order_status:
        for order in logistic_log.logistics.orders.all():
            order.order_status = order_status
            if order_status == 'delivered':
                order.delivery_date = logistic_log.updated_time
            order.save()
            OrderTracker.objects.create(orders_id=order.id,status=order_status)
        for returns in logistic_log.logistics.returns.all():
            returns.return_status = order_status
            if order_status == 'picked_up':
                returns.order_id.order_status = 'return_in_process'
                returns.order_id.save()
            if order_status == 'on_the_way_to_you':
                returns.order_id.order_status = 'return_shipped'
                returns.order_id.save()
            if order_status == 'delivered':
                returns.order_id.order_status = 'returned'
                returns.order_id.save()
                returns.delivery_date = logistic_log.updated_time
            returns.save()


def track_order(log_id, partner):
    # pdb.set_trace()
    data = {
            "status1" : "Data Not Available",
            "status2" : "Data Not Available"
        }
    if not partner in ['AR','DL']:
        return data
    waybill = LogisticsLog.objects.get(id=log_id).waybill
    if partner == 'AR':
        shipments = [str(waybill)]
        client_info = {
            "UserName": settings.ARAMEX_USERNAME,
            "Password": settings.ARAMEX_PASSWORD,
            "Version": settings.ARAMEX_VERSION,
            "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
            "AccountPin": settings.ARAMEX_ACCOUNT_PIN,
            "AccountEntity": settings.ARAMEX_ACCOUNT_ENTITY,
            "AccountCountryCode": settings.ARAMEX_ACCOUNT_COUNTRY_CODE
        }
        
        aramex_transaction = {
            "Reference1": "",
            "Reference2": "",
            "Reference3": "",
            "Reference4": "",
            "Reference5": ""
        }
        url = settings.ARAMEX_BASE_URL + settings.ARAMEX_UPDATE_URL
        data = {"ClientInfo": client_info, "GetLastTrackingUpdateOnly":
                True, "Shipments": shipments, "Transaction": aramex_transaction}
        headers = {
            'Content-Type': 'application/json', 'Accept': 'application/json'}
        resp = requests.post(url, data=json.dumps(data), headers=headers)
        json_response = resp.json()

        # print json_response,'json_response'
        # print json_response["TrackingResults"][0]["Value"][0]["ProblemCode"]
        # print json_response["TrackingResults"][0]["Value"][0]["UpdateCode"]
        ar = AramexStatus.objects.get(status_code=json_response["TrackingResults"][0]["Value"][0]["UpdateCode"])
        data = {
            "status1" : ar.code_description,
            "status2" : ar.customer_description
        }
        print data
    else:
        url = settings.DELHIVERY_BASE_URL + "/api/packages/json/?token=" + settings.DELHIVERY_API_KEY + "&waybill=" + str(waybill)
        resp_product = requests.get(url)
        json_response = resp_product.json()    
        data = {
            "status1" : json_response['ShipmentData'][0]['Shipment']['Status']['Status'],
            "status2" : json_response['ShipmentData'][0]['Shipment']['Status']['Instructions']
        }          
    return data


# Updates delhivery status.
@periodic_task(run_every=(crontab(minute=30, hour=[3, 14])))
def delhivery_normal_update():
    # pdb.set_trace()
    track_list = LogisticsLog.objects.filter(track=True, partner='DL')
    track_dict = [(x.waybill, x) for x in track_list]
    track_dict = dict(track_dict)
    if track_list:

        waybill_list = list(track_list.values_list('waybill', flat=True))
        waybill_string = ",".join(waybill_list)

        url = settings.DELHIVERY_BASE_URL + "/api/packages/json/?token=" + \
            settings.DELHIVERY_API_KEY + "&waybill=" + waybill_string
        try:

            resp_product = requests.get(url)
            json_response = resp_product.json()

            for json in json_response["ShipmentData"]:
                # waybill_index = waybill_list.index(json["Shipment"]["AWB"])

                logistic_log = track_dict.get(json["Shipment"]["AWB"])
                status = get_status_value(
                    'DL', json["Shipment"]["Status"]["Status"])

                if logistic_log.log_status < status:
                    status_serializer = LogisticsLogSerializer(data={"log_status": status,
                                                                     "updated_time": json["Shipment"]["Status"]["StatusDateTime"]})
                    if status_serializer.is_valid():
                        status_serializer.update(
                            logistic_log, status_serializer.validated_data)
                    update_order_status(status, logistic_log)

        except Exception as e:
            print e

"""
Updates parcelled status.
"""
@periodic_task(run_every=(crontab(minute=30, hour=[3, 14])))
def parcelled_update():
    # pdb.set_trace()
    track_list = LogisticsLog.objects.filter(track=True, partner='PR')
    track_dict = [(x.waybill, x) for x in track_list]
    track_dict = dict(track_dict)
    if track_list:
        waybill_list = list(track_list.values_list('waybill', flat=True))
        waybill_string = ",".join(waybill_list)

        url = settings.PARCELLED_BASE_URL + 'bookings/' + waybill_string
        headers = {'parcelled-api-key': settings.PARCELLED_API_KEY,
                   'Content-Type': 'application/json'}
        resp = requests.get(url, headers=headers)
        json_response = resp.json()

        if (json_response['meta']['code'] == 200 or json_response['meta']['code'] == 201):
            for json in json_response['data']['bookings']:
                logistic_log = track_dict.get(json["booking_id"])

                status = get_status_value('PR', json["status"])
                if 'shipments' in json and logistic_log.log_status < status:

                    status_serializer = LogisticsLogSerializer(
                        data={"status": status, "updated_time": json['shipments']['trackings']['timestamp']})
                    if status_serializer.is_valid():
                        status_serializer.update(
                            logistic_log, status_serializer.validated_data)

                    update_order_status(status, logistic_log)

"""
Updates Aramex status.
"""
@periodic_task(run_every=(crontab(minute=30, hour=[3, 14])))
def aramex_update():
    from datetime import datetime, timedelta
    import json

    track_list = LogisticsLog.objects.filter(track=True, partner='AR')
    track_dict = [(x.waybill, x) for x in track_list]
    track_dict = dict(track_dict)
    print track_list,'tracklist 1734'
    if track_list:
        waybill_list = track_list.values_list('waybill', flat=True)
        waybill_list = map(str, waybill_list)
        client_info = {
            "UserName": settings.ARAMEX_USERNAME,
            "Password": settings.ARAMEX_PASSWORD,
            "Version": settings.ARAMEX_VERSION,
            "AccountNumber": settings.ARAMEX_ACCOUNT_NUMBER,
            "AccountPin": settings.ARAMEX_ACCOUNT_PIN,
            "AccountEntity": settings.ARAMEX_ACCOUNT_ENTITY,
            "AccountCountryCode": settings.ARAMEX_ACCOUNT_COUNTRY_CODE
        }
        shipments = waybill_list
        aramex_transaction = {
            "Reference1": "",
            "Reference2": "",
            "Reference3": "",
            "Reference4": "",
            "Reference5": ""
        }
        url = settings.ARAMEX_BASE_URL + settings.ARAMEX_UPDATE_URL
        data = {"ClientInfo": client_info, "GetLastTrackingUpdateOnly":
                True, "Shipments": shipments, "Transaction": aramex_transaction}
        headers = {
            'Content-Type': 'application/json', 'Accept': 'application/json'}
        resp = requests.post(url, data=json.dumps(data), headers=headers)
        json_response = resp.json()
        print json_response,'json resp 1761'
        aramex_status = list(AramexStatus.objects.all().values(
            'status_code', 'problem_code', 'logistic_status'))

        if json_response and not json_response['HasErrors']:
            for json in json_response['TrackingResults']:
                logistic_log = track_dict.get(json["Key"])
                try:
                    status = filter(lambda key: key['status_code'] == json['Value'][0]['UpdateCode'] and key[
                                    'problem_code'] == (json['Value'][0]['ProblemCode'] or None), aramex_status)[0]['logistic_status']

                    if logistic_log.log_status < status:
                        update_datetime_list = re.findall(
                            r'\d+', json['Value'][0]['UpdateDateTime'])
                        updated_time = datetime.fromtimestamp(
                            int(update_datetime_list[0])/1000.0)

                        status_serializer = LogisticsLogSerializer(
                            data={"status": status, "updated_time": updated_time})
                        if status_serializer.is_valid():
                            status_serializer.update(
                                logistic_log, status_serializer.validated_data)

                        update_order_status(status, logistic_log)
                except:
                    print 'Status not found'

"""
Excel file created for outgoing consignments from zapyle on each day.
"""
def create_outgoing_excel(directory):
    if not os.path.isfile(directory + "/zapyle_outgoing.xlsx"):
        int_wb = Workbook()
        int_ws = int_wb.active
        # put headers for xcel
        int_ws['A1'] = 'Delivery Partner'
        int_ws['B1'] = 'Order Reference'
        int_ws['C1'] = 'Order Id'
        int_ws['D1'] = 'Return Id'
        int_ws['E1'] = 'Filename'
        int_ws['F1'] = 'Product Id'
        int_ws['G1'] = 'Bill filename'
        int_wb.save(directory + "/zapyle_outgoing.xlsx")

"""
Updates the outgoing excel file.
"""
def update_outgoing_excel(excel_location, logistic_dict):
    int_wb = load_workbook(excel_location)
    int_ws = int_wb.active
    int_ini = int_ws.max_row + 1
    int_ws['A' + str(int_ini)] = logistic_dict['partner']
    int_ws['B' + str(int_ini)] = logistic_dict['order_number']
    int_ws['C' + str(int_ini)] = logistic_dict['order_ids']
    int_ws['D' + str(int_ini)] = logistic_dict['return_ids']
    int_ws['E' + str(int_ini)] = logistic_dict['filename']
    int_ws['F' + str(int_ini)] = logistic_dict['product_ids']
    int_ws['G' + str(int_ini)] = logistic_dict['bill_filename']
    int_ini += 1
    # print str(int_ini) + " excel"
    int_wb.save(excel_location)

"""
Saves a zapyle bill according to order.
"""
def save_zap_bill(order, directory):
    zapemail = ZapEmail()
    codDecider, zapcashDecider, bill_cod, paid_amt = 0, 0, 0, 0
    zapwallet_used = order.zapwallet_used

    bill_total = float(order.final_price * order.quantity)
    bill_net = order.final_payable_price() + order.zapwallet_used
    order_date = order.placed_at.date().strftime('%d-%m-%Y')
    if order.zapwallet_used:
        zapcashDecider = 1
    if order.transaction.payment_mode == 'cod':
        # for zapyle Bill
        codDecider = 1
        bill_cod = order.final_payable_price()
    else:
        paid_amt = order.final_payable_price()
    if order.consignor.user.user_type.name == "store_front":
        userTypeDecider = 0
    else:
        userTypeDecider = 1
    consignee = order.transaction.consignee

    zapyle_bill_template_data = {'order_number': order.order_number, 'to_name': consignee.name, 'to_add': consignee.address, 'to_city': consignee.city,
                                 'to_state': consignee.state.name, 'to_pincode': consignee.pincode, 'to_phone': consignee.phone, 'package_product': order.ordered_product.title,
                                 'qty': order.quantity, 'package_price': order.final_price, 'total': bill_total, 'shipping': order.shipping_charge,
                                 'net': bill_net, 'cod_amt': bill_cod, 'userTypeDecider': userTypeDecider, 'codDecider': codDecider, 'zapcashDecider': zapcashDecider,
                                 'zapwallet_used': zapwallet_used, 'order_date': order_date, 'paid_amt': paid_amt}
    zapyle_bill_html_body = render_to_string(
        "zapyle_bill.html", zapyle_bill_template_data)

    filename = directory + "/ZAPBILL_" + order.order_number + ".html"
    if not os.path.exists(os.path.dirname(filename)):
        try:
            os.makedirs(os.path.dirname(filename))
        except OSError as exc: # Guard against race condition
            print "error in creating folder"

    Html_file = open(
        directory + "/ZAPBILL_" + order.order_number + '.html', "w+")
    Html_file.write(zapyle_bill_html_body)
    Html_file.close()
    
    if order.zapwallet_used:
        email_body = "Greetings,\n\nPFA the bill with order number" + order.order_number + ".\n\nThe amount of zapcash used for this transaction is:" + \
            str(order.zapwallet_used) + \
            ".\nWe will send you the service invoice at the time of payouts. Please refer to this bill for reference."
    else:
        email_body = "Greetings,\n\nPFA the bill with order number" + order.order_number + \
            ".\nWe will send you the service invoice at the time of payouts. Please refer to this bill for reference."
    
    attachment = directory + "/ZAPBILL_" + order.order_number + ".html"
    attachment_name = "BILL_" + order.order_number + ".html"
    zapemail.send_email_attachment("Order BILL - " + order.order_number, settings.FROM_EMAIL, settings.ZAPYLE_BILL_EMAIL, email_body=email_body, attachment=attachment, attachment_name=attachment_name)


    pdf_url = directory + "/ZAPBILL_" + \
        order.order_number + ".pdf"
    if not os.path.exists(os.path.dirname(pdf_url)):
        try:
            os.makedirs(os.path.dirname(pdf_url))
        except OSError as exc: # Guard against race condition
            print "error in creating folder"


    config = pdfkit.configuration(wkhtmltopdf=settings.WKHTMLTOPDF_PATH)
    # pdfkit.from_string(zapyle_bill_html_body, pdf_url, configuration=config)
    pdfkit.from_string(zapyle_bill_html_body, pdf_url)

    attachment = directory + "/ZAPBILL_" + order.order_number + ".pdf"
    attachment_name = "BILL_" + order.order_number + ".pdf"
    zapemail.send_email_attachment("Order BILL - " + order.order_number, settings.FROM_EMAIL, settings.ZAPYLE_BILL_EMAIL, email_body=email_body, attachment=attachment, attachment_name=attachment_name)                

    return "ZAPBILL_" + order.order_number + ".pdf"

"""
When the product reaches zapyle.
"""
# @periodic_task(run_every=(crontab(minute=0, hour=6)))
def pickup_update(logistic_id=None, delivery_type='normal'):
    from datetime import datetime, timedelta, time
    # packaging material related queries
    # pdb.set_trace()
    zapemail = ZapEmail()
    if logistic_id:
        reached_zapyle_list = LogisticsLog.objects.filter(logistics_id=logistic_id,
                                                          pickup=True)#track=True, log_status=4
        zapexclusive_inhouse = Logistics.objects.filter(
            id=logistic_id, triggered_pickup=False, pickup_partner=None, orders__isnull=False).distinct()
        print str(logistic_id) + 'logistic id'
    else:
        reached_zapyle_list = LogisticsLog.objects.filter(
            track=True, pickup=True, log_status=4)

        zapexclusive_inhouse = Logistics.objects.filter(
            triggered_pickup=False, pickup_partner=None, orders__isnull=False).distinct()
    aramex_logistic_ref = reached_zapyle_list.filter(
        logistics__product_delivery_partner='AR').values_list('logistics__id', flat=True)
    aramex_logistic_ref_inhouse = zapexclusive_inhouse.filter(
        product_delivery_partner='AR').values_list('id', flat=True)
    aramex_logistic_ref = list(
        chain(aramex_logistic_ref, aramex_logistic_ref_inhouse))
    if aramex_logistic_ref:

        seller_add = {
            "Line1": "Zapyle",
            "Line2": "CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli",
            "Line3": "",
            "City": "Bangalore",
            "StateOrProvinceCode": "Karnataka",
            "PostCode": "560016",
            "CountryCode": "IN"
        }
        seller_contact = {
            "PersonName": "Zapyle",
            "CompanyName": "Zapyle",
            "PhoneNumber1": "07337880282",
            "PhoneNumber2": "",
            "FaxNumber": "",
            "CellPhone": "9663596594",
            "EmailAddress": "hello@zapyle.com",
            "Type": ""
        }

        aramex_logistic_ref = str(aramex_logistic_ref)

        '''if timezone.now().hour <= 2 and timezone.now().minute <= 30:
            aramex_pickup_datetime = timezone.now() + \
                timezone.timedelta(hours=5)
            aramex_pickup_readytime = timezone.now(
            ) + timezone.timedelta(hours=5, minutes=5)
            aramex_pickup_lasttime = timezone.now(
            ) + timezone.timedelta(hours=8, minutes=5)
            aramex_pickup_closetime = aramex_pickup_lasttime + \
                timezone.timedelta(days=1)
        elif timezone.now().hour >= 11:
            aramex_pickup_datetime = datetime.combine(timezone.now().date() +
                                                      timezone.timedelta(days=1), time(12))
            aramex_pickup_readytime = datetime.combine(timezone.now().date() +
                                                       timezone.timedelta(days=1), time(12, 5))
            aramex_pickup_lasttime = aramex_pickup_readytime + \
                timezone.timedelta(hours=7, minutes=30)
            aramex_pickup_closetime = aramex_pickup_lasttime + \
                timezone.timedelta(days=1)
        else:
            aramex_pickup_datetime = datetime.combine(timezone.now().date(),time(12))
            aramex_pickup_readytime = datetime.combine(timezone.now().date(),time(12, 5))
            # aramex_pickup_datetime = timezone.now() + \
            #     timezone.timedelta(hours=3)
            # aramex_pickup_readytime = timezone.now(
            # ) + timezone.timedelta(hours=3, minutes=5)
            aramex_pickup_lasttime = datetime.combine(timezone.now().date(),
                                                      time(14))
            aramex_pickup_closetime = aramex_pickup_lasttime + \
                timezone.timedelta(days=1)'''

        if timezone.now().hour <= 6 and timezone.now().minute <= 30: # 0.00am to 12 pm
            aramex_pickup_datetime = datetime.combine(timezone.now().date(), time(12, 5))
            aramex_pickup_readytime = datetime.combine(timezone.now().date(), time(12, 10))
            aramex_pickup_lasttime = datetime.combine(timezone.now().date(), time(14))
            aramex_pickup_closetime = aramex_pickup_lasttime + timezone.timedelta(days=1)
        elif timezone.now().hour >= 11 and timezone.now().minute >= 30: # 5.00pm to 12am
            aramex_pickup_datetime = datetime.combine(timezone.now().date() +
                                                      timezone.timedelta(days=1), time(12))
            aramex_pickup_readytime = datetime.combine(timezone.now().date() +
                                                       timezone.timedelta(days=1), time(12, 5))
            aramex_pickup_lasttime = aramex_pickup_readytime + \
                timezone.timedelta(hours=7, minutes=30)
            aramex_pickup_closetime = aramex_pickup_lasttime + \
                timezone.timedelta(days=1)
        else:
            aramex_pickup_datetime = timezone.now() + timezone.timedelta(hours=5, minutes=35)
            aramex_pickup_readytime = timezone.now() + timezone.timedelta(hours=5, minutes=40)
            aramex_pickup_lasttime = timezone.now() + timezone.timedelta(hours=8, minutes=5)
            aramex_pickup_closetime = aramex_pickup_lasttime + timezone.timedelta(days=1)
        print aramex_pickup_datetime,'aramex_pickup_datetime'
        print aramex_pickup_readytime,'aramex_pickup_readytime'
        print aramex_pickup_lasttime,'aramex_pickup_lasttime'
        print aramex_pickup_closetime,'aramex_pickup_closetime'

        aramex_pickup_request_extra = {
            'aramex_pickup_datetime': str(int((aramex_pickup_datetime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
            'aramex_pickup_readytime': str(int((aramex_pickup_readytime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
            'aramex_pickup_lasttime': str(int((aramex_pickup_lasttime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
            'aramex_pickup_closetime': str(int((aramex_pickup_closetime.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),

            'shipment_qty': len(aramex_logistic_ref)
        }
        print timezone.now(),'now'
        # pdb.set_trace()
        # print aramex_pickup_request_extra,'aramex_pickup_request_extra'
        json_response = aramex_pickup_create(
            aramex_logistic_ref, len(aramex_logistic_ref), seller_add, seller_contact, "P", aramex_pickup_request_extra)
        print json_response,'json res 6373'
        if json_response and not json_response['HasErrors']:
            aramex_pickup_guid = json_response['ProcessedPickup']['GUID']
            aramex_pickup_extra = json.dumps(json_response['ProcessedPickup'])

        if timezone.now().hour <= 2 and timezone.now().minute <= 30:
            aramex_shipment_time = timezone.now() + timezone.timedelta(hours=5)
            aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

        elif timezone.now().hour >= 11:
            aramex_shipment_time = datetime.combine(timezone.now().date() +
                                                    timezone.timedelta(days=1), time(6, 5))
            aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

        else:
            aramex_shipment_time = timezone.now() + timezone.timedelta(hours=3, minutes=5)
            aramex_duedate_time = aramex_shipment_time + timezone.timedelta(days=1)

        aramex_extra = {
            'shipping_time': str(int((aramex_shipment_time.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
            'duedate_time': str(int((aramex_duedate_time.replace(tzinfo=None) - datetime(1970, 1, 1)).total_seconds()*1000)),
        }

    zapexclusive_inhouse_id_list = zapexclusive_inhouse.values_list(
        'id', flat=True)

    pickup_logistics_list = reached_zapyle_list.values_list(
        'logistics', flat=True)
    pickup_logistics_list = list(chain(
        pickup_logistics_list, zapexclusive_inhouse_id_list))

    #//FIX THIS
    pincode_logistics = Logistics.objects.filter(
        id__in=pickup_logistics_list)
    consignor_pincode = pincode_logistics.values_list(
        'consignor__pincode', flat=True)
    consignee_pincode = pincode_logistics.values_list(
        'consignee__pincode', flat=True)
    pincode_list = []
    pincode_list = list(set(list(chain(consignor_pincode, consignee_pincode))))

    pincodes_object_list = DelhiveryPincode.objects.filter(
        pincode__in=pincode_list).values_list('pincode', 'coc_code')

    pincode_coc_finder = dict(pincodes_object_list)
    #//FIX THIS

    reached_zapyle_list = list(chain(
        reached_zapyle_list, zapexclusive_inhouse))

    if reached_zapyle_list:

        todays_date = timezone.now().date().strftime('%d-%m-%Y')
        directory = settings.HOME_FOLDER + "/operations/" + todays_date

        after_two_days = timezone.now().date() + timezone.timedelta(2)
        after_two_days = after_two_days.strftime('%d-%m-%Y')
        future_directory = settings.HOME_FOLDER + \
            "/operations/" + after_two_days

        create_delhivery_excel(directory)
        # FOR INTERNAL EXCEL SHEET
        create_outgoing_excel(directory)

    delhivery_logistics_finder = {}

    for log in reached_zapyle_list:
        try:
            logistic = log.logistics
        except:
            logistic = log

        if (logistic.returns.all() and not (None in logistic.returns.all().values_list('product_verification', flat=True))) or (logistic.orders.all()
                                                                                                                                and not (None in logistic.orders.all().values_list('product_verification', flat=True))):
            # if not (None in logistic.orders.all().values_list('product_verification', flat=True) or
            # None in
            # logistic.returns.all().values_list('product_verification',
            # flat=True)):
            if delivery_type == 'normal':  #delivery to buyer
                consignee = logistic.consignee #buyer
                consignor = logistic.consignor #seller
            else: # delivery to seller (return)
                consignor = logistic.consignee 
                consignee = logistic.consignor
            products_desc = ""
            total_amount = 0
            cod_amt = 0
            order_number = ""
            cod_decider = 0

            # ######pdb.set_trace()
            product_category = ""
            qty = 0

            product_ids = ''
            zapyle_bill = ''
            internal_order_ids = ''
            internal_return_ids = ''
            if logistic.returns.all():
                returns_decider = True
                waybill_zapyle = 'ZAPRETDEL'
                delhivery_service = 'DL_ret_delivery'
                zapyle_service = 'ZP_ret_delivery'

                for returns in logistic.returns.all():

                    products_desc = products_desc + " " + str(returns.order_id.quantity) + " " + returns.order_id.ordered_product.color + \
                        " " + returns.order_id.ordered_product.size + returns.order_id.product.brand.brand + " " + \
                        returns.order_id.ordered_product.product_category + ","
                    products_desc = products_desc.replace(
                        products_desc[-1:], '.')
                    total_amount += (
                        returns.order_id.ordered_product.total_price * returns.order_id.quantity)
                    order_number += returns.order_id.order_number + ", "
                    product_category += returns.order_id.ordered_product.product_category + \
                        ", "
                    qty += returns.order_id.quantity
                    product_ids += str(returns.order_id.product.id) + ", "
                    internal_return_ids += str(returns.id) + ", "
                # order_number = order_number[:-2]
                internal_return_ids = internal_return_ids[:-2]

            else:
                # SAVE

                returns_decider = False
                waybill_zapyle = 'ZAPDEL'
                delhivery_service = 'DL_delivery'
                zapyle_service = 'ZP_delivery'
                if logistic.consignor.user.user_type.name == "store_front":
                    userTypeDecider = 0
                else:
                    userTypeDecider = 1

                for order in logistic.orders.all():

                    internal_order_ids += str(order.id) + ", "
                    product_ids += str(order.product.id) + ", "
                    products_desc = products_desc + " " + str(order.quantity) + " " + order.ordered_product.color + " " + order.ordered_product.size + \
                        " " + order.ordered_product.brand + " " + \
                        order.ordered_product.product_category + ","
                    products_desc = products_desc.replace(
                        products_desc[-1:], '.')
                    if order.transaction.payment_mode == 'cod':
                        # for packing slip
                        cod_decider = 1
                        cod_amt += order.final_payable_price()
                    # SAVE A BILL
                    try:
                        zapyle_bill += save_zap_bill(order, directory) + ", "

                    except Exception as e:
                        print 'bill not saved'
                        print '%s (%s)' % (e.message, type(e))

                    total_amount += order.total_price()
                    order_number += order.order_number + ", "
                    product_category += order.ordered_product.product_category + \
                        ", "
                    qty += order.quantity

                # order_number = order_number[:-2]
                internal_order_ids = internal_order_ids[:-2]
                zapyle_bill = zapyle_bill[:-2]

            product_ids = product_ids[:-2]

            order_number = order_number[:-2]
            product_category = product_category[
                :len(product_category) - 2] + product_category[-2:].replace(product_category[-2:], ' ')

            # DELHIVERY DELIVERY STARTS
            # CREATE WITH WAYBILL, SAVE IN EXCEL, ######CALL PACKING SLIP AND SAVE HTML IN SAME DAY'S FOLDER
            # ######pdb.set_trace()
            if not logistic.product_delivery_partner:
                # handle zap_inhouse
                # Email to Buyer acknowledging the receipt of product. Issue zapcash for Buyer
                # print "Acknowledge buyer, send email to ganesh if needed"

                # template_data = {
                #     'buyer_name': consignor.user.get_full_name(), 'product_desc': products_desc}
                # # html_body = render_to_string(
                # #     "buyer_after_returns.html", template_data)

                # html = settings.BUYER_AFTER_RETURN
                # html_body = render_to_string(
                #     html['html'], template_data)

                # zapemail.send_email_alternative(html[
                #     'subject'], settings.FROM_EMAIL, consignor.user.email, html_body)

                try:
                    returns_serializer = ReturnSerializer(
                        data={'delivery_date': log.updated_time})
                    if returns_serializer.is_valid():
                        for returns in logistic.returns.all():
                            returns_serializer.update(
                                returns, returns_serializer.validated_data)

                            # INCREASE NUMBER OF QUANTITY
                            #????????????
                            if settings.CELERY_USE:
                                increase_quantity_with_size.delay(
                                    returns.order_id.ordered_product.size, returns.order_id.product.id, returns.order_id.quantity)
                            else:
                                increase_quantity_with_size(
                                    returns.order_id.ordered_product.size, returns.order_id.product.id, returns.order_id.quantity)
                            # no_of_prod = returns.order_id.product.product_count.get(
                            #     size=returns.order_id.size)
                            # no_of_prod.quantity += returns.order_id.quantity
                            # no_of_prod.save()

                    # Do put product back to sale and increase size6of12

                    log.track = False
                    log.save()

                    # print '12'
                except:
                    print "Not Increased"
            elif logistic.product_delivery_partner == 'AR':

                seller_shipment = {
                    'address': seller_add,
                    'contact': seller_contact
                }
                consignee_shipment = {
                    'address': {
                        "Line1": consignee.address,
                        "Line2": consignee.address2 or "",
                        "Line3": "",
                        "City": consignee.city,
                        "StateOrProvinceCode": consignee.state.name,
                        "PostCode": consignee.pincode,
                        "CountryCode": "IN"
                    },
                    'contact': {
                        "PersonName": consignee.name,
                        "CompanyName": consignee.name,
                        "PhoneNumber1": consignee.phone,
                        "PhoneNumber2": "",
                        "FaxNumber": "",
                        "CellPhone": consignee.phone,
                        "EmailAddress": consignee.user.email,
                        "Type": ""
                    }
                }
                aramex_ret_ref = "Returns - " + \
                    str(logistic.returns.all().values_list(
                        'id', flat=True)) if returns_decider else ""
                aramex_extra['total_amount'] = total_amount
                aramex_extra['cod'] = True if cod_decider else False
                aramex_extra['cod_amt'] = cod_amt

                shipment_resp = aramex_shipment_create(
                    logistic, seller_shipment, consignee_shipment, aramex_pickup_guid, "P", aramex_extra, ret=aramex_ret_ref)
                print shipment_resp,' shipment_resp'
                if shipment_resp and not shipment_resp['HasErrors']:
                    status = 0
                    shipment_data = shipment_resp['Shipments'][0]
                    delivery_data = {'logistics': logistic.id, 'waybill': shipment_data['ID'],
                                     'log_status': status, 'track': True, 'whole_response': json.dumps(shipment_resp),
                                     'logistics_ref': str(logistic.id), 'partner': 'AR', 'pickup': False,
                                     'returns': returns_decider, 'extra': json.dumps(json_response['ProcessedPickup'])}
                    if not delivery_type == 'normal':
                        delivery_data['rejected'] = True
                    print 'success11'
                    delivery_log_serializer = LogisticsLogSerializer(
                        data=delivery_data)
                    if delivery_log_serializer.is_valid():
                        delivery_log_serializer.save()
                        # update_order_status(2,log)
                        try:
                            log.track = False
                            log.save()
                        except:
                            log.triggered_pickup = True
                            log.save()
                        if not delivery_type == 'normal':
                            for order in log.orders.all():
                                OrderTracker.objects.create(orders_id=order.id,status="return_in_process")
                                order.order_status = 'return_in_process'
                                order.save()

                    if not logistic.triggered_pickup:
                        logistic.triggered_pickup = True
                        logistic.save()
                    aramex_pdf_url = directory + "/delivery_slip_aramex_" + \
                        shipment_data['ID'] + ".pdf"
                    if settings.CELERY_USE:
                        save_aramex_pdf.delay(
                            shipment_data['ShipmentLabel']['LabelURL'], aramex_pdf_url)
                    else:
                        save_aramex_pdf(
                            shipment_data['ShipmentLabel']['LabelURL'], aramex_pdf_url)

                    logistic_dict = {'partner': 'Aramex', 'order_number': order_number, 'order_ids': internal_order_ids, 'return_ids': internal_return_ids,
                                     'filename': "delivery_slip_aramex_" + shipment_data['ID'] + ".pdf", 'product_ids': product_ids, 'bill_filename': zapyle_bill}
                    update_outgoing_excel(
                        directory + "/zapyle_outgoing.xlsx", logistic_dict)

            elif logistic.product_delivery_partner == 'DL':
                waybill_json = delhivery_create_waybill()
                pickup_location = {'add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore',
                                   'country': 'India', 'name': settings.DELHIVERY_PICKUP_NAME, 'phone': '07337880282', 'pin': '560016'}
                if cod_amt:
                    shipments = {'waybill': waybill_json, 'name': consignee.name, 'order': logistic.id, 'products_desc': products_desc,
                                 'order_date': timezone.now().isoformat(), 'payment_mode': 'COD', 'total_amount': total_amount, 'cod_amount': cod_amt,
                                 'add': consignee.address, 'city': consignee.city, 'state': consignee.state.name, 'pin': consignee.pincode,
                                 'phone': consignee.phone, 'weight': "",  'product_quantity': 1, 'return_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli',
                                 'return_city': 'Bangalore', 'return_name': 'Zapyle', 'return_phone': '07337880282', 'return_pin': '560016',
                                 'return_state': 'Karnataka', 'seller_name': consignor.user.get_full_name()}
                else:
                    shipments = {'waybill': waybill_json, 'name': consignee.name, 'order': logistic.id, 'products_desc': products_desc,
                                 'order_date': timezone.now().isoformat(), 'payment_mode': 'Pre-paid', 'total_amount': total_amount, 'add': consignee.address,
                                 'city': consignee.city, 'state': consignee.state.name, 'pin': consignee.pincode, 'phone': consignee.phone,
                                 'weight': "",  'product_quantity': 1, 'return_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'return_city': 'Bangalore',
                                 'return_name': 'Zapyle', 'return_phone': '07337880282', 'return_pin': '560016', 'return_state': 'Karnataka',
                                 'seller_name': consignor.user.get_full_name()}

                delivery_json_response, delivery_content = delhivery_create(
                    shipments, pickup_location, waybill_json)

                if (delivery_json_response) and 'success' in delivery_json_response and delivery_json_response['success'] is True:
                    status = get_status_value(
                        'DL', delivery_json_response["packages"][0]["status"])
                    response_data = {'logistics': logistic.id, 'waybill': waybill_json, 'log_status': status, 'track': True,
                                     'whole_response': delivery_content, 'logistics_ref': delivery_json_response["upload_wbn"], 'pickup': False,
                                     'partner': 'DL', 'returns': returns_decider}
                    if not delivery_type == 'normal':
                        delivery_data['rejected'] = True
                    # print 'success22'
                    log_serializer = LogisticsLogSerializer(data=response_data)
                    if log_serializer.is_valid():
                        log_serializer.save()
                        if not delivery_type == 'normal':
                            for order in log.orders.all():
                                OrderTracker.objects.create(orders_id=order.id,status="return_in_process")
                                order.order_status = 'return_in_process'
                                order.save()
                        # update_order_status(2,log)
                        try:
                            log.track = False
                            log.save()
                        except:
                            log.triggered_pickup = True
                            log.save()

                        logistic_dict = {'address': 'Zapyle, CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore',
                                         'state': 'Karnataka', 'pincode': '560016', 'phone': '07337880282'}
                        update_delhivery_excel(
                            directory+"/pickup.xlsx", logistic_dict)
                        if not logistic.triggered_pickup:
                            logistic.triggered_pickup = True
                            logistic.save()

                        delhivery_logistics_finder.update(
                            {waybill_json: logistic})

                else:
                    if 'error' in delivery_json_response:
                        error_message = delivery_json_response['error']
                    else:
                        error_message = delivery_json_response["rmk"]
                    delivery_response_data = {'logistics_ref': delivery_json_response[
                        "upload_wbn"], 'error_message': error_message, 'returns': returns_decider}
                    delivery_log_serializer = LogisticsLogSerializer(
                        data=delivery_response_data)
                    if delivery_log_serializer.is_valid():
                        delivery_log_serializer.save()

            # DELHIVERY DELIVERY ENDS

            # ZAPYLE DELIVERY STARTS
            # CREATE LOGISTICS LOG , SAVE HTML
            elif logistic.product_delivery_partner == 'ZP':

                delivery_data = {'logistics': logistic.id, 'waybill': waybill_zapyle + str(logistic.id), 'log_status': 2, 'track': True,
                                 'whole_response': 'Schedule a pickup and change to Manifested once Amzad is informed.', 'logistics_ref': str(logistic.id),
                                 'partner': 'ZP', 'pickup': False, 'returns': returns_decider}
                if not delivery_type == 'normal':
                        delivery_data['rejected'] = True
                # print 'success11'
                delivery_log_serializer = LogisticsLogSerializer(
                    data=delivery_data)
                if delivery_log_serializer.is_valid():
                    delivery_log_serializer.save()
                    if not delivery_type == 'normal':
                        for order in log.orders.all():
                            OrderTracker.objects.create(orders_id=order.id,status="return_in_process")
                            order.order_status = 'return_in_process'
                            order.save()
                    update_order_status(2,log) #amzad pickedup
                    try:
                        log.track = False
                        log.save()
                    except:
                        log.triggered_pickup = True
                        log.save()

                # ########pdb.set_trace()
                if logistic.consignor.user.user_type.name == "store_front":
                    userTypeDecider = 0
                else:
                    userTypeDecider = 1

                packing_slip_template_data = {'todays_date': todays_date, 'order_number': order_number, 'from_name': 'Zapyle',
                                              'from_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'from_city': 'Bangalore', 'from_state': 'Karnataka',
                                              'from_pincode': '560016', 'from_phone': '07337880282', 'to_name': logistic.consignee.name, 'to_add': logistic.consignee.address,
                                              'to_city': logistic.consignee.city, 'to_state': logistic.consignee.state.name, 'to_pincode': logistic.consignee.pincode,
                                              'to_phone': logistic.consignee.phone, 'package_product': products_desc, 'package_price': str(total_amount),
                                              'userTypeDecider': userTypeDecider, 'cod_decider': cod_decider, 'cod_amt': cod_amt}
                packing_slip_html_body = render_to_string(
                    "zapyle_delivery_slip.html", packing_slip_template_data)

                Html_file = open(
                    directory + "/" + zapyle_service + "_" + str(logistic.id) + '.html', "w")
                Html_file.write(packing_slip_html_body)
                Html_file.close()

                if not logistic.triggered_pickup:
                    logistic.triggered_pickup = True
                    logistic.save()
                logistic_dict = {'partner': 'Zapyle', 'order_number': order_number, 'order_ids': internal_order_ids, 'return_ids': internal_return_ids,
                                 'filename': zapyle_service + "_" + str(logistic.id) + '.html', 'product_ids': product_ids, 'bill_filename': zapyle_bill}
                update_outgoing_excel(
                    directory + "/zapyle_outgoing.xlsx", logistic_dict)

            # ZAPYLE DELIVERY ENDS

            # PARCELLED DELIVERY STARTS

            elif logistic.product_delivery_partner == 'PR':

                pr_logistic_id = logistic.id

                pr_booking_item = {}
                if cod_amt:
                    pr_booking_item.update({'is_cod': True})
                    pr_booking_item.update({'value': cod_amt})
                pr_booking_item.update({'category': product_category})
                pr_booking_item.update({'qty': qty})

                pr_booking_item = [pr_booking_item]
                pr_consignee = {'address': consignee.address, 'city': consignee.city, 'country': consignee.country,
                                'name': consignee.name, 'phone': consignee.phone, 'pincode': consignee.pincode, 'state': consignee.state.name}
                pr_consignor = {'address': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'city': 'Bangalore', 'country': 'India',
                                'name': 'Zapyle', 'phone': '07337880282', 'pincode': '560016', 'state': 'Karnataka', 'email': 'shafi@zapyle.com'}
                # pr_consignor = {'address':consignor.address, 'city':consignor.city, 'country':consignor.country, 'name':consignor.name, 'phone':consignor.phone, 'pincode':consignor.pincode, 'state':consignor.state.name, 'email':consignor.user.email}

                json_response = parcelled_create(
                    str(pr_logistic_id), pr_booking_item, pr_consignee, pr_consignor)

                if (json_response['meta']['code'] == 200 or json_response['meta']['code'] == 201):
                    status = get_status_value(
                        'PR', json_response['data']['bookings'][0]['status'])
                    delivery_data = {'logistics': logistic.id, 'waybill': json_response['data']['bookings'][0]['booking_id'], 'log_status': status,
                                     'track': True, 'whole_response': json.dumps(json_response), 'logistics_ref': json_response['data']['bookings'][0]['parent_order']['order_id'],
                                     'partner': 'PR', 'pickup': False, 'returns': returns_decider}
                    if not delivery_type == 'normal':
                        delivery_data['rejected'] = True
                    print 'success11'
                    delivery_log_serializer = LogisticsLogSerializer(
                        data=delivery_data)
                    if delivery_log_serializer.is_valid():
                        delivery_log_serializer.save()
                        if not delivery_type == 'normal':
                            for order in log.orders.all():
                                OrderTracker.objects.create(orders_id=order.id,status="return_in_process")
                                order.order_status = 'return_in_process'
                                order.save()
                        # update_order_status(2,delivery_log_serializer)
                        try:
                            log.track = False
                            log.save()
                        except:
                            log.triggered_pickup = True
                            log.save()

                    if not logistic.triggered_pickup:
                        logistic.triggered_pickup = True
                        logistic.save()
                    logistic_dict = {'partner': 'Parcelled', 'order_number': order_number, 'order_ids': internal_order_ids, 'return_ids': internal_return_ids,
                                     'filename': "", 'product_ids': product_ids, 'bill_filename': zapyle_bill}
                    update_outgoing_excel(
                        directory + "/zapyle_outgoing.xlsx", logistic_dict)

        else:
            if not (logistic.orders.all() or logistic.returns.all()):
                try:
                    log.track = False
                    log.save()
                except:
                    log.triggered_pickup = True
                    log.save()
    # PARCELLED DELIVERY ENDS
    waybill_keys = delhivery_logistics_finder.keys()
    waybill_str = ",".join(waybill_keys)

    packing_slip = delhivery_packing_slip(waybill_str)
    if packing_slip and 'packages' in packing_slip:
        for packages in packing_slip['packages']:

            logistics_data = delhivery_logistics_finder.get(packages['wbn'])

            order_number = ""
            internal_order_ids = ''
            internal_return_ids = ''
            product_ids = ''
            zapyle_bill = ''

            if logistics_data.returns.all():
                for ret in logistics_data.returns.all():
                    order_number += ret.order_id.order_number + ", "
                    product_ids += str(ret.order_id.product.id) + ", "
                    internal_return_ids += str(ret.id) + ", "
                internal_return_ids = internal_return_ids[:-2]
                delhivery_service = 'DL_ret_delivery'

            else:
                for single_order in logistics_data.orders.all():
                    order_number += single_order.order_number + ", "
                    internal_order_ids += str(single_order.id) + ", "
                    product_ids += str(single_order.product.id) + ", "
                    zapyle_bill += "ZAPBILL_" + order.order_number + ".pdf, "
                internal_order_ids = internal_order_ids[:-2]
                delhivery_service = 'DL_delivery'
                zapyle_bill = zapyle_bill[:-2]

            product_ids = product_ids[:-2]

            order_number = order_number[:-2]
            coc_code = pincode_coc_finder.get(
                logistics_data.consignee.pincode, '')

            oid_barc = packages['oid_barcode']
            wbn_barc = packages['barcode']
            oid_barcode = decode_base64(oid_barc)
            wbn_barcode = decode_base64(wbn_barc)
            oid_filename = directory + "/" + delhivery_service + \
                "_oid_bar_" + packages['wbn'] + ".png"
            wbn_filename = directory + "/" + delhivery_service + \
                "_wbn_bar_" + packages['wbn'] + ".png"
            delhivery_logo_full = packages['delhivery_logo']
            delhivery_logo_raw = delhivery_logo_full.split(":")[1]
            delhivery_logo = "https://track.delhivery.com" + delhivery_logo_raw
            with open(oid_filename, 'wb') as f:
                f.write(oid_barcode)
            f.close()
            with open(wbn_filename, 'wb') as f:
                f.write(wbn_barcode)
            f.close()

            # ########pdb.set_trace()
            if logistics_data.consignor.user.user_type.name == "store_front":
                userTypeDecider = 0
            else:
                userTypeDecider = 1
            if packages['cod']:
                cod_decider = 1
                cod_amt = packages['cod']
            else:
                cod_decider = 0
                cod_amt = 0

            waybill_barcode = settings.CURRENT_DOMAIN + "/logmedia/" + todays_date + \
                "/" + delhivery_service + "_wbn_bar_" + \
                packages['wbn'] + ".png"
            order_barcode = settings.CURRENT_DOMAIN + "/logmedia/" + todays_date + \
                "/" + delhivery_service + "_oid_bar_" + \
                packages['wbn'] + ".png"

            packing_slip_template_data = {'delhivery_logo': delhivery_logo, 'waybill_barcode': waybill_barcode, 'order_barcode': order_barcode,
                                          'order_number': order_number, 'coc_code': coc_code, 'from_name': 'Zapyle', 'from_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli',
                                          'from_city': 'Bangalore', 'from_state': 'Karnataka', 'from_pincode': '560016', 'from_phone': '07337880282', 'to_name': logistics_data.consignee.name,
                                          'to_add': logistics_data.consignee.address, 'to_city': logistics_data.consignee.city, 'to_state': logistics_data.consignee.state.name,
                                          'to_pincode': logistics_data.consignee.pincode, 'to_phone': logistics_data.consignee.phone, 'package_product': packages['prd'],
                                          'package_price': str(packages['rs']), 'userTypeDecider': userTypeDecider, 'cod_decider': cod_decider, 'cod_amt': cod_amt}
            packing_slip_html_body = render_to_string(
                "delhivery_packing_slip.html", packing_slip_template_data)

            Html_file = open(settings.HOME_FOLDER + "/operations/" +
                             todays_date + "/" + delhivery_service + "_" + packages['wbn'] + '.html', "w")
            Html_file.write(packing_slip_html_body)
            Html_file.close()

            logistic_dict = {'partner': 'Delhivery', 'order_number': order_number, 'order_ids': internal_order_ids, 'return_ids': internal_return_ids,
                             'filename': delhivery_service + "_" + packages['wbn'] + '.html', 'product_ids': product_ids, 'bill_filename': zapyle_bill}
            update_outgoing_excel(
                directory + "/zapyle_outgoing.xlsx", logistic_dict)

    #####

"""
Shipment creation for products with zapyle and products which have reached zapyle from seller.
"""
@periodic_task(run_every=(crontab(minute=30, hour=[6, 15])))
def product_update():
    zapemail = ZapEmail()
    delivery_logistics_list = LogisticsLog.objects.filter(
        track=True, pickup=False, log_status__in=[4, 6])
    if delivery_logistics_list:
        seller_after_return_msg = "Hi Zapyler! Your item has been returned to you and it's going back as a listing on your Zapyle Closet. Keep the listing going strong!"

        sms_msg = "Hey style diva! We hope you loved your stylish parcel tied with love. In the case you didn't, you have 24 hours to return it with a valid reason. Toodles!"
        zapsms = ZapSms()
    for log in delivery_logistics_list:

        # check if product is order or return and update accordingly
        # ########pdb.set_trace()
        if log.returns == False:
            if log.log_status == 4 and not log.logistics.orders.filter(product_verification__in=['rejected_shipped, rejected']):
                products = get_products_template_data(
                    log.logistics.orders.all())
                # SEND FEEDBACK FORM TO BUYER
                template_data = {
                    'buyer_name': log.logistics.consignee.user.get_full_name(), 'products': products}
                html_body = render_to_string(
                    "buyer_feedback.html", template_data)

                html = settings.BUYER_FEEDBACK
                html_body = render_to_string(
                    html['html'], template_data)
                zapemail.send_email_alternative(
                    html['subject'], settings.FROM_EMAIL, log.logistics.consignee.user.email, html_body)

                try:
                    zapsms.send_sms(
                        log.logistics.consignee.user.phone_number, sms_msg)
                except Exception as e:
                    print e

        else:
            if log.log_status == 4 and not log.logistics.returns.all()[0].order_id.ordered_product.with_zapyle:

                if log.logistics.consignee.user.user_type.name == 'zap_exclusive':
                    userTypeDecider = 0
                    sellerdata = log.logistics.returns.all()[0].order_id.product.zapexclusiveuserdata_set.all()[
                        0]
                    seller_name = sellerdata.full_name
                    seller_email = sellerdata.email
                    seller_phone = sellerdata.phone_number
                else:
                    seller_name = log.logistics.consignee.user.get_full_name()
                    seller_email = log.logistics.consignee.user.email
                    seller_phone = log.logistics.consignee.user.phone_number

                products_desc = ""
                for ret in log.logistics.returns.all():
                    products_desc = products_desc + " " + str(ret.order_id.quantity) + " " + ret.order_id.ordered_product.color + " " + \
                        ret.order_id.ordered_product.size + ret.order_id.ordered_product.brand + \
                        " " + \
                        ret.order_id.ordered_product.product_category + ","
                    products_desc = products_desc.replace(
                        products_desc[-1:], '.')
                    # INCREASE NUMBER OF QUANTITY
                    if settings.CELERY_USE:
                        increase_quantity_with_size.delay(
                            ret.order_id.ordered_product.size, ret.order_id.product.id, ret.order_id.quantity)
                    else:
                        increase_quantity_with_size(
                            ret.order_id.ordered_product.size, ret.order_id.product.id, ret.order_id.quantity)
                    # no_of_prod = ret.order_id.product.product_count.get(
                    #     size=ret.order_id.size)
                    # no_of_prod.quantity += ret.order_id.quantity
                    # no_of_prod.save()

                # SEND BUYER THE EMAIL
                buyer_template_data = {'buyer_name': log.logistics.consignor.user.get_full_name(
                ), 'product_desc': products_desc}

                html = settings.BUYER_AFTER_RETURN
                html_body = render_to_string(
                    html['html'], buyer_template_data)
                if not log.logistics.orders.filter(product_verification__in=['rejected_shipped, rejected']):
                    zapemail.send_email_alternative(html[
                        'subject'], settings.FROM_EMAIL, log.logistics.consignor.user.email, html_body)
                # SEND EMAIL TO SELLER
                profile_link = settings.CURRENT_DOMAIN + \
                    "/#/profile/" + str(log.logistics.consignee.user.id)
                seller_template_data = {'seller_name': seller_name, 'profile_link': profile_link,
                                        'zap_username': log.logistics.consignee.user.zap_username}

                html = settings.SELLER_AFTER_RETURN
                html_body = render_to_string(
                    html['html'], seller_template_data)
                if not log.logistics.orders.filter(product_verification__in=['rejected_shipped, rejected']):
                    zapemail.send_email_alternative(html[
                        'subject'], settings.FROM_EMAIL, seller_email, html_body)
                    try:
                        zapsms.send_sms(seller_phone, seller_after_return_msg)
                    except Exception as e:
                        print e

            elif log.log_status == 6:
                pickup_logistics = log.logistics.logistics_log.filter(
                    pickup=True)[0]
                pickup_logistics.track = True
                pickup_logistics.save()

            # Do put product back to sale and increase size6of12
        log.track = False
        log.save()
        logistic_log_serializer = LogisticsLogSerializer(data={'track': False})
        if logistic_log_serializer.is_valid():
            logistic_log_serializer.update(
                log, logistic_log_serializer.validated_data)

#@periodic_task(run_every=(crontab(minute=0, hour='*/3')))


# def update_delhivery_status():

#     delhivery_normal_update()
#     delhivery_packaging_update()
#     delhivery_product_update()


#@periodic_task(run_every=(crontab(minute=30, hour=14)))
# @periodic_task(run_every=(crontab(minute=30, hour=0)))
# def zapcash_approval():
#     returns = Return.objects.filter(
#         delivery_date__isnull=False, approved_zapcash=False).distinct()
#     returns_serializer = ReturnSerializer(data={'approved_zapcash': True})
#     if returns_serializer.is_valid():
#         for ret in returns:
#             if (timezone.now() - ret.delivery_date).days > 2:
#                 returns_serializer.update(
#                     ret, returns_serializer.validated_data)


# #"cqoos6l2u2"
# #"R05H4Q7GB1XX3SWQ98IF"

# # @periodic_task(run_every=(crontab(minute=0, hour=2)))
# def issue_zapcash():
#     returns = Return.objects.filter(credited=False, approved_zapcash=True)

#     url = 'https://sandboxcoupons.citruspay.com/cms_api.asmx/IssueCoupons'
#     headers = {'Content-Type': 'application/json'}

#     for ret in returns:
#         # email = ret.consignor.user.email
#         # phone = ret.consignor.user.phone_number
#         # amt = ret.value
#         # data = {"CampaignCode": "ZAPCASH", "UserList": [{"Email": email, "Mobile": phone, "Amount": amt, "Type": "F"}],"PartnerID": "za61q0w33q " ,"Password": "GS3K0UGZOG8WQ6KQIG61" }
#         try:
#             # resp = requests.post(url,data=json.dumps(data),headers=headers)
#             # json_response = resp.json()
#             # send the buyer email
#             ret.consignor.user.issue_zap_wallet(
# ret.value, mode='2', purpose='Credited for Return', return_id=ret.id)

#             returns_serializer = ReturnSerializer(data={'credited': True})
#             if returns_serializer.is_valid():
#                 print returns_serializer.validated_data
#                 returns_serializer.update(
#                     ret, returns_serializer.validated_data)
#                 # Email to buyer for returned zapcash
#                 zapemail = ZapEmail()
#                 html = settings.ZAPCASH_RETURN_HTML
#                 email_vars = {
#                     'user': ret.consignor.user.get_full_name() or ret.consignor.user.zap_username,
#                 }

#                 html_body = render_to_string(
#                     html['html'], email_vars)

#                 zapemail.send_email_alternative(html[
#                     'subject'], settings.FROM_EMAIL, ret.consignor.user.email, html_body)

#                 # zapemail.send_email(html['html'], html[
#                 #                     'subject'], email_vars, settings.FROM_EMAIL, ret.consignor.user.email)
#             # print returns_serializer.errors
#             # return Response(json_response)
#         except Exception as e:
#             print e
#             # return Response({'error': e})

"""
Rejected orders/returns that needs to be shipped back to buyer/seller.
"""
def rejected_products(order_id=None):
    orders_list, orders_id_list, all_pincode = [], [], []
    # if order_number:
    #     orders_list = Order.objects.filter(product_verification='rejected_shipped')
    # else:
    if order_id:
        orders_list = Order.objects.filter(
            id=order_id)
    else:
        orders_list = Order.objects.filter(
            product_verification='rejected')
    returns_queryset, all_return_delhivery_object, all_return_pincode = [], [], []
    # if ret_ids:
    returns_queryset = Return.objects.filter(product_verification='rejected')

    # else:
    #     returns_queryset = Return.objects.filter(
    #         order_id__order_status='return_requested', approved=True, self_return=False)
    if orders_list:
        all_pincode = orders_list.values_list('consignor__pincode', flat=True)
    # print returns_queryset, 'returns_queryset'
    if returns_queryset:
        all_return_pincode = returns_queryset.values_list(
            'consignor__pincode', flat=True)
        all_pincode = list(chain(all_pincode, all_return_pincode))
        all_return_delhivery_object = DelhiveryPincode.objects.filter(
            pincode__in=all_return_pincode).values_list('pincode', flat=True)
        pincode_finder = dict(all_return_delhivery_object)

    # if all_pincode:
    #     parcelled_json_resp = parcelled_pickup_check(
    #         all_pincode)
    #     if parcelled_json_resp["meta"]["code"] == 200:
    #         parcelled_pickup_finder = parcelled_json_resp["data"]
    #     else:
    #         parcelled_pickup_finder = {}

    evaluated_order_dict = {}
    # pdb.set_trace()
    for order in orders_list:

        # evaluated_dict = {}
        # if order.ordered_product.with_zapyle:
        #     order.product_verification = 'approved'
        #     order.save()
        pickup_logistics = None
        if 'bangalore' in order.consignor.city.lower() or 'bengaluru' in order.consignor.city.lower():
            delivery_logistics = 'ZP'
        elif settings.ARAMEX_SERVICE:
            delivery_logistics = 'AR'
        # elif parcelled_pickup_finder.get(order.consignor.pincode, False) and settings.PARCELLED_SERVICE:
        #     pickup_logistics = 'PR'
        elif settings.DELHIVERY_SERVICE:
            delivery_logistics = 'DL'

        # if 'bangalore' in order.transaction.consignee.city.lower() or 'bengaluru' in order.transaction.consignee.city.lower():
        #     delivery_logistics = 'ZP'
        # elif settings.ARAMEX_SERVICE:
        #     delivery_logistics = 'AR'
        # elif settings.DELHIVERY_SERVICE:
        #     delivery_logistics = 'DL'

        evaluated_dict = {order.id: {
            'pickup_logistics': pickup_logistics, 'delivery_logistics': delivery_logistics}}
        evaluated_order_dict.update(evaluated_dict)
        data = {'orders': [order.id], 'consignee': order.transaction.consignee.id, 'consignor': order.consignor.id, 'pickup_partner': pickup_logistics, 'product_delivery_partner': delivery_logistics}
        log_serializer = LogisticsSerializer(data=data)
        if log_serializer.is_valid():
            log_serializer.save()
            # order.product_verification='return_requested'
            order.order_status = 'return_requested'
            order.save()
            OrderTracker.objects.create(orders_id=order.id,status="return_requested")

    # evaluated_returns_dict = {}
    # for returns in returns_queryset:

    #     # Do evaluation of self return
    #     repl = pincode_finder.get(returns.consignor.pincode, 'N')
    #     # if repl == 'N':
    #     #     # returns_id_list.remove(returns.id)
    #     #     returns_serializer = ReturnSerializer(data={'self_return': True})
    #     #     if returns_serializer.is_valid():
    #     #         returns_serializer.update(
    #     #             returns, returns_serializer.validated_data)
    #     #         pickup_logistics = 'SR'

    #     # else:
            
    #     pickup_logistics = None

    #     if returns.consignor.city.lower() == 'bangalore' or returns.consignor.city.lower() == 'bengaluru':
    #         delivery_logistics = 'ZP'
    #     elif settings.ARAMEX_SERVICE:
    #         delivery_logistics = 'AR'
    #     # elif parcelled_pickup_finder.get(returns.consignor.pincode, False) and settings.PARCELLED_SERVICE:
    #     #     pickup_logistics = 'PR'
    #     elif settings.DELHIVERY_SERVICE:
    #         delivery_logistics = 'DL'

    #     # 
    #     returns_dict = {returns.id: {
    #         'pickup_logistics': pickup_logistics, 'delivery_logistics': delivery_logistics}}
    #     evaluated_returns_dict.update(returns_dict)
    # pdb.set_trace()
    # if evaluated_order_dict:
    #     print 'call optimize_order_logistics'
    #     if 0 and settings.CELERY_USE:
    #         optimize_order_logistics.delay(evaluated_order_dict)
    #     else:
    #         optimize_order_logistics(
    #             evaluated_order_dict)
    #     Order.objects.filter(
    #         product_verification='rejected').update(product_verification='return_requested')
    #     print 'end optimize_order_logistics'
    # if settings.CELERY_USE:
    #     print 'call optimize_return_logistics'
    #     if ret_ids:
    #         optimize_return_logistics.delay(evaluated_returns_dict)
    #     else:
    #         optimize_return_logistics(
    #             evaluated_returns_dict)
    #     Return.objects.filter(product_verification='rejected').update(product_verification='rejected_shipped')
    #     print 'end optimize_return_logistics'
    print evaluated_order_dict,'evaluated_order_dict'
    return evaluated_order_dict



"""
Account details manipulation
"""
def decode_base64(data):

    data = data[data.find(",") + 1:]

    missing_padding = 4 - len(data) % 4
    if missing_padding:
        data += ('=' * missing_padding)
    return base64.b64decode(data)

"""
Logistics log creations for self returns.
"""
@periodic_task(run_every=(crontab(minute=30, hour=6)))
def self_returns():
    # ########pdb.set_trace()
    zapemail = ZapEmail()

    return_ids = Logistics.objects.filter(triggered_logistics=False, returns__isnull=False,
                                          returns__self_return=True, approved=True).distinct().values_list('returns', flat=True)

    returns_product_list = Return.objects.filter(id__in=return_ids)

    todays_date = timezone.now().date().strftime('%d-%m-%Y')
    directory = settings.HOME_FOLDER + "/operations/" + todays_date
    if not os.path.exists(directory):
        os.makedirs(directory)

    for ret in returns_product_list:

                # returns_logisticslog_data = returns_logisticslog_finder.get(packages['wbn'])

        return_email = ret.consignor.user.email
        print ret.order_id.order_number, ' order_number'
        order_number = str(ret.order_id.order_number)
        # for single_order in ret.order_id.all():
        #     order_number += single_order.order_number + ", "
        # order_number = order_number[:-2]

        products_desc = str(ret.order_id.quantity) + " " + ret.order_id.ordered_product.color.name + " " + ret.order_id.ordered_product.size + \
            " " + ret.order_id.ordered_product.brand + " " + \
            ret.order_id.ordered_product.product_category + "."

        self_returns_template_data = {'order_number': order_number, 'from_name': ret.consignor.name, 'from_add': ret.consignor.address,
                                      'from_city': ret.consignor.city, 'from_state': ret.consignor.state.name, 'from_pincode': ret.consignor.pincode,
                                      'from_phone': ret.consignor.phone, 'to_name': 'Zapyle', 'to_add': 'CoWorks, RMZ Infinity, No.3, Opp. Gopalan Signature Mall, Old Madras Road, Bennigana Halli', 'to_city': 'Bangalore',
                                      'to_state': 'Karnataka', 'to_pincode': '560016', 'to_phone': '07337880282', 'package_product': products_desc,
                                      'package_price': str(ret.order_id.total_price())}
        self_returns_html_body = render_to_string(
            "self_returns.html", self_returns_template_data)

        # save html

        Html_file = open(directory + "/self_ret_" + str(ret.id) + '.html', "w")
        Html_file.write(self_returns_html_body)
        Html_file.close()

        # convert html to pdf
        pdf_url = directory + "/self_ret_" + str(ret.id) + ".pdf"
        config = pdfkit.configuration(wkhtmltopdf=settings.WKHTMLTOPDF_PATH)
        pdfkit.from_string(
            self_returns_html_body, pdf_url, configuration=config)
        # send diff kind of return email to buyer
        email_body = "Greetings,\n\nPlease go through the instructions given in the pdf attached."
        attachment = directory + "/self_ret_" + str(ret.id) + ".pdf"
        attachment_name = "self_return_" + str(ret.id) + ".pdf"
        zapemail.send_email_attachment("Returns Document - Zapyle", settings.FROM_EMAIL, return_email,
                                       email_body=email_body, attachment=attachment, attachment_name=attachment_name)

        logistic = ret.return_logistic.all()[0]
        logistic.triggered_logistics = True
        logistic.save()
        # except:
        #     print "Not Sent"


# @periodic_task(run_every=(crontab(minute=0, hour=1)))
"""
Payout status check.
"""
def check_payout_initiation():
    Q1 = Q(payout_status__isnull=True)
    Q2 = Q(order_status='delivered')
    past_date = timezone.now() - timezone.timedelta(2)
    Q3 = Q(delivery_date__lte=past_date)
    orders = Order.objects.filter(Q1 & Q2 & Q3)
    for order in orders:
        if not order.order_status in ['cancelled', 'return_requested', 'return_in_process', 'returned']:
            order.payout_status = 'can_initiate_payout'
            order.save()

"""
Returns plain account number from an encrypted account number.
"""
def decryptAcc(acc_number):
    try:
        obj = DES.new(settings.ENCKEY, DES.MODE_ECB)
        base64_user_acc_no = acc_number
        cipher_user_acc_no = base64.b64decode(base64_user_acc_no)
        user_acc_no_dummy = obj.decrypt(cipher_user_acc_no)
        user_acc_no = re.sub('[Z]', '', user_acc_no_dummy)
        # len_user_acc_no = len(user_acc_no)
        # user_acc = user_acc_no[:2] + (len_user_acc_no-4)*'X' + user_acc_no[len_user_acc_no-2:]
        return user_acc_no
    except Exception as e:
        return e

#@periodic_task(run_every=(crontab(minute=0, hour=9)))

# def base10toN(num,n):
# return ((num == 0) and  "0" ) or ( base10toN(num // n, n).strip("0") +
# "0123456789abcdefghijklmnopqrstuvwxyz"[:n][num % n])

"""
Emails are sent for payouts.
"""
@periodic_task(run_every=(crontab(minute=30, hour=2)))
def payouts(order_id=[]):
    # pdb.set_trace()
    zapemail = ZapEmail()
    num2wrd = inflect.engine()
    orders = Order.objects.filter(payout_status='can_initiate_payout')
    todays_date = timezone.now().date().strftime('%d-%m-%Y')
    directory = settings.HOME_FOLDER + "/operations/" + todays_date
    if not os.path.exists(directory):
        os.makedirs(directory)

    if timezone.now().date().month > 3:
        financial_year = str(
            timezone.now().date().year - 1)+'-'+str(timezone.now().date().year)[2:]
    else:
        financial_year = str(
            timezone.now().date().year - 2)+'-'+str(timezone.now().date().year - 1)[2:]
    return_result_list = []

    for order in orders:
        try:
            seller_name = ''
            return_result_dict = {}
            invoice_num = "INV" + order.order_number[2:]
            if order.product.user.user_type.name in ('zap_exclusive', 'zap_dummy'):
                userdata = order.product.zapexclusiveuserdata_set.all()[
                    0]
                seller_name = userdata.full_name
                seller_email = userdata.email
            else:
                userdata = order.product.user.user_data
                seller_name = order.product.user.get_full_name()
                seller_email = order.product.user.email

            if userdata.account_holder and userdata.account_number and userdata.ifsc_code:

                actual_seller_price = (
                    order.final_price * order.quantity)
                seller_amt = (
                    (100-order.ordered_product.percentage_commission)/100.0)*actual_seller_price
                template_data = {'seller': seller_name, 'amt': seller_amt}
                zapyle_amt = float(actual_seller_price) - float(seller_amt)
                activity = "Commission @ " + \
                    str(order.ordered_product.percentage_commission) + \
                    "% on Rs. " + str(actual_seller_price)+"/-"
                amt_words = num2wrd.number_to_words(zapyle_amt)

                invoice_template_data = {'seller_name': seller_name, 'address': order.consignor.address, 'city': order.consignor.city,
                                         'state': order.consignor.state.name, 'pincode': order.consignor.pincode, 'country': order.consignor.country,
                                         'invoice': invoice_num, 'date': todays_date, 'order_ref': order.order_number, 'activity': activity, 'amt': str(zapyle_amt),
                                         'amt_words': amt_words, 'financial_year': financial_year}
                invoice_html_body = render_to_string(
                    "service_invoice.html", invoice_template_data)

                pdf_url = directory + "/service_invoice_" + \
                    invoice_num + ".pdf"
                config = pdfkit.configuration(
                    wkhtmltopdf=settings.WKHTMLTOPDF_PATH)
                pdfkit.from_string(
                    invoice_html_body, pdf_url, configuration=config)
                attachment = directory + \
                    "/service_invoice_" + invoice_num + ".pdf"
                attachment_name = "service_invoice_" + invoice_num + ".pdf"

                html = settings.SELLER_PAYOUT
                html_body = render_to_string(
                    html['html'], template_data)

                email_body = "Greetings,\n\nPFA the service invoice for " + order.transaction.payment_mode + " order. Please make a payment to below given seller.\n\nAccount Holder Name: " + \
                    userdata.account_holder+"\nAccount Number: " + \
                    decryptAcc(userdata.account_number)+"\nIFSC No.: " + \
                    userdata.ifsc_code+"\nAmount: "+str(seller_amt)
                attachment = settings.HOME_FOLDER + "/operations/" + \
                    todays_date + "/service_invoice_" + invoice_num + ".pdf"
                attachment_name = "service_invoice_" + invoice_num + ".pdf"
                zapemail.send_email_attachment(order.transaction.payment_mode+" Order - Service Invoice: "+invoice_num, settings.FROM_EMAIL, [
                                               'shafi@zapyle.com', 'accounts@zapyle.com', 'likhita@zapyle.com'], email_body=email_body, attachment=attachment, attachment_name=attachment_name)
                #,'accounts@zapyle.com', 'likhita@zapyle.com'
                zapemail.send_email_alternative(
                    html['subject'], settings.FROM_EMAIL, seller_email, html_body, attachment, attachment_name)

                order.payout_status = 'paid_out'
                order.save()
                if order.final_payable_price - seller_amt < 0:
                    zapyle_cut = 0
                else:
                    zapyle_cut = order.final_payable_price - seller_amt
                payout_serializer = PayoutSerializer(
                    data={'seller': order.product.user.id, 'order': order.id, 'total_value': actual_seller_price, 'seller_cut': seller_amt, 'zapyle_cut': zapyle_cut})
                if payout_serializer.is_valid():
                    payout_record = payout_serializer.save()

                return_result_dict['id'] = order.id
                return_result_dict['mode'] = order.transaction.payment_mode
                return_result_dict['status'] = True
                # except:
                #     print "not sent to Ganesh"
            else:
                # SEND EMAIL TO SELLER THAT WE DONT HAVE THEIR ACC INFO
                # print "SEND EMAIL"
                # ###pdb.set_trace()
                if seller_name:
                    seller_template_data = {
                        'seller_name': seller_name}
                    html = settings.SELLER_BANK_INFO
                    html_body = render_to_string(
                        html['html'], seller_template_data)

                    zapemail.send_email_alternative(
                        html['subject'], settings.FROM_EMAIL, seller_email, html_body)

                    return_result_dict['id'] = order.id
                    return_result_dict['mode'] = order.transaction.payment_mode
                    return_result_dict['status'] = False

                else:
                    body = "Check the order id" + str(order.id)
                    zapemail.send_email_attachment(
                        "Zapyle - Seller's Email is not available for Payout!", settings.FROM_EMAIL, 'shafi@zapyle.com', email_body=body)

                    return_result_dict['id'] = order.id
                    return_result_dict['mode'] = order.transaction.payment_mode
                    return_result_dict['status'] = False
        except:
            body = "Check the order id. Got Error. Not Style Tribute" + \
                str(order.id)
            zapemail.send_email_attachment(
                "Zapyle - Seller's Email is not available for Payout!", settings.FROM_EMAIL, 'shafi@zapyle.com', email_body=body)
            return_result_dict['id'] = order.id
            return_result_dict['mode'] = order.transaction.payment_mode
            return_result_dict['status'] = False
        return_result_list.append(return_result_dict)
    return return_result_list
