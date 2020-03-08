import json
import sys, os
import pdb
import base64
from django.conf import settings
from openpyxl import load_workbook, Workbook
from zap_apps.zap_catalogue.models import *
from zap_apps.zap_catalogue.product_serializer import ZapProductSerializer, ProductImageSerializer, NumberOfProductSrlzr
from zap_apps.zapuser.models import *
from zap_apps.zap_analytics.models import *
from zap_apps.cart.models import *
from zap_apps.order.models import *
import urllib2
import re
from shutil import copyfile, make_archive
import datetime
import time
import zipfile
from django.http import HttpResponse, HttpResponseRedirect

class DesignerAnalytics:

    def __init__(self):
        pass

    def all_designers(self, start_date, end_date):
        users = ZapUser.objects.filter(user_type__name='designer')

        timestamp = int(time.time())
        cwd = os.getcwd() + '/zap_media/analytics_data/'
        folder_name = cwd + str(timestamp)

        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        for user in users:
            products = ApprovedProduct.objects.filter(user=user.id)

            designer_name = user.username
            designer_id = str(user.id)
            file_name = folder_name + "/" + designer_name + designer_id + ".xlsx"

            copyfile('zap_apps/zap_analytics/seller_analytics/templat.xlsx', file_name)

            wb = load_workbook(filename=file_name)
            ws = wb['Data']

            i = 2

            while i <= len(products):
                for product in products:
                    product_id = product.id
                    product_name = product.title
                    product_price = product.listing_price
                    product_discount = product.discount
                    product_category = product.product_category.parent.name
                    product_subcategory = product.product_category.name
                    total_product_impressions = ImpressionAnalytics.objects.filter(product=product.id).count()
                    total_product_views = ProductAnalytics.objects.filter(product=product.id).count()
                    total_added_to_cart = Item.objects.filter(product=product.id).count()
                    total_orders = Order.objects.filter(product=product.id).count()
                    period_product_impressions = ImpressionAnalytics.objects.filter(product=product.id,
                                                                                    time__range=[start_date, end_date]).count()
                    period_product_views = ProductAnalytics.objects.filter(product=product.id,
                                                                           time__range=[start_date, end_date]).count()
                    period_added_to_cart = Item.objects.filter(product=product.id,
                                                               added_at__range=[start_date, end_date]).count()
                    period_orders = Order.objects.filter(product=product.id,
                                                         placed_at__range=[start_date, end_date]).count()

                    ws['A' + str(i)] = product_id
                    ws['B' + str(i)] = product_name
                    ws['C' + str(i)] = total_product_impressions
                    ws['D' + str(i)] = total_product_views
                    ws['E' + str(i)] = total_added_to_cart
                    ws['F' + str(i)] = total_orders
                    ws['G' + str(i)] = period_product_impressions
                    ws['H' + str(i)] = period_product_views
                    ws['I' + str(i)] = period_added_to_cart
                    ws['J' + str(i)] = period_orders
                    ws['K' + str(i)] = product_price
                    ws['L' + str(i)] = product_discount
                    ws['M' + str(i)] = product_subcategory
                    ws['N' + str(i)] = product_category

                    i += 1

            wb.save(file_name)

        zip_file = folder_name

        make_archive(zip_file, "zip", folder_name)

        return "https://www.zapyle.com/zapmedia/analytics_data/%s.zip" % (str(timestamp), )

    def individual_designer(self, user_id, start_date, end_date):
        user = ZapUser.objects.get(id=user_id)

        products = ApprovedProduct.objects.filter(user=user.id)

        designer_name = user.username
        designer_id = str(user.id)

        cwd = os.getcwd() + '/zap_media/analytics_data/'
        file_name = cwd + "/" + designer_name + designer_id + str(start_date) + str(end_date) + ".xlsx"

        copyfile('zap_apps/zap_analytics/seller_analytics/templat.xlsx', file_name)

        wb = load_workbook(filename=file_name)
        ws = wb['Data']

        i = 2

        while i <= len(products):
            for product in products:
                product_id = product.id
                product_name = product.title
                product_price = product.listing_price
                product_discount = product.discount
                product_category = product.product_category.parent.name
                product_subcategory = product.product_category.name
                total_product_impressions = ImpressionAnalytics.objects.filter(product=product.id).count()
                total_product_views = ProductAnalytics.objects.filter(product=product.id).count()
                total_added_to_cart = Item.objects.filter(product=product.id).count()
                total_orders = Order.objects.filter(product=product.id).count()
                period_product_impressions = ImpressionAnalytics.objects.filter(product=product.id,
                                                                                time__range=[start_date,
                                                                                             end_date]).count()
                period_product_views = ProductAnalytics.objects.filter(product=product.id,
                                                                       time__range=[start_date, end_date]).count()
                period_added_to_cart = Item.objects.filter(product=product.id,
                                                           added_at__range=[start_date, end_date]).count()
                period_orders = Order.objects.filter(product=product.id,
                                                     placed_at__range=[start_date, end_date]).count()

                ws['A' + str(i)] = product_id
                ws['B' + str(i)] = product_name
                ws['C' + str(i)] = total_product_impressions
                ws['D' + str(i)] = total_product_views
                ws['E' + str(i)] = total_added_to_cart
                ws['F' + str(i)] = total_orders
                ws['G' + str(i)] = period_product_impressions
                ws['H' + str(i)] = period_product_views
                ws['I' + str(i)] = period_added_to_cart
                ws['J' + str(i)] = period_orders
                ws['K' + str(i)] = product_price
                ws['L' + str(i)] = product_discount
                ws['M' + str(i)] = product_subcategory
                ws['N' + str(i)] = product_category

                i += 1

        wb.save(file_name)

        return "https://www.zapyle.com/zapmedia/analytics_data/" + designer_name + designer_id + str(start_date) + str(end_date) + ".xlsx"