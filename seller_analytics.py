import json
# import pdb
import sys, os
# sys.path.append("/path/to/project")
import django

try:
    env = sys.argv[1]
except IndexError:
    raise Exception("Please provide settings [local/ staging/ prod]")

os.environ["DJANGO_SETTINGS_MODULE"] = "zapyle_new.settings.{}".format(env)
django.setup()
import pdb
import base64
from django.conf import settings
from openpyxl import load_workbook, Workbook
from zap_apps.zap_catalogue.models import *
from zap_apps.zap_catalogue.product_serializer import ZapProductSerializer, ProductImageSerializer, NumberOfProductSrlzr
from zap_apps.zapuser.models import *
from zap_apps.zap_analytics.models import *
import urllib2
import re
from shutil import copyfile
import datetime

users = ZapUser.objects.filter(user_type__name='designer')

import time
timestamp = int(time.time())
folder_name = "seller_analytics/" + str(timestamp)

if not os.path.exists(folder_name):
    os.makedirs(folder_name)

for user in users:
    products = ApprovedProduct.objects.filter(user=user.id)

    designer_name = user.username
    designer_id = str(user.id)
    file_name = folder_name + "/" + designer_name + designer_id + ".xlsx"

    copyfile('template.xlsx', file_name)

    wb = load_workbook(filename=file_name)
    ws = wb['Data']

    i = 2

    while i <= len(products):
        for product in products:
            product_id = product.id
            product_name = product.title
            product_price = product.listing_price
            product_discount = product.discount
            product_category = product.product_category.parent.name
            product_subcategory = product.product_category.name
            product_impressions = None
            product_views = None
            added_to_cart = None
            orders = None

            ws['A' + str(i)] = product_id
            ws['B' + str(i)] = product_name
            ws['C' + str(i)] = product_impressions
            ws['D' + str(i)] = product_views
            ws['E' + str(i)] = added_to_cart
            ws['F' + str(i)] = orders
            ws['G' + str(i)] = product_price
            ws['H' + str(i)] = product_discount
            ws['I' + str(i)] = product_subcategory
            ws['J' + str(i)] = product_category

            i += 1

    wb.save(file_name)
