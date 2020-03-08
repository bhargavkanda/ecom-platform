import json
#import pdb
import sys,os
# sys.path.append("/path/to/project")
import django
try:
    env = sys.argv[1]
except IndexError:
    raise Exception("Please provide settings [local/ staging/ prod]")

os.environ["DJANGO_SETTINGS_MODULE"] = "zapyle_new.settings.{}".format(env)
django.setup()
import pdb
from django.conf import settings
from openpyxl import load_workbook, Workbook
from zap_apps.marketing.models import NotificationTracker
from dateutil import parser
from django.utils import timezone



# wb = load_workbook(filename=sys.argv[2], data_only=True)
# ws = wb['Sheet1'] # ws is now an IterableWorksheet
# data = {}

wb = Workbook()
ws = wb.active
# put headers for xcel
ws['A1'] = 'Notification Text'
ws['B1'] = 'Action'
ws['C1'] = 'User'
ws['D1'] = 'Device'
ws['E1'] = 'Sent Time'
ws['F1'] = 'Opened Time'
todays_date = timezone.now().date().strftime('%d-%m-%Y')
ini = 2
try:
	notifs = NotificationTracker.objects.filter(sent_time__range=(parser.parse(sys.argv[2]),parser.parse(sys.argv[3]) + timezone.timedelta(days=1)))

	for notif in notifs:
	    ws['A'+str(ini)] = notif.notif.text
	    ws['B'+str(ini)] = notif.notif.action.action_type
	    ws['C'+str(ini)] = notif.user.zap_username or notif.user.get_full_name()
	    ws['D'+str(ini)] = notif.user.logged_device.name
	    ws['E'+str(ini)] = notif.sent_time.date().strftime('%d-%m-%Y')
	    ws['F'+str(ini)] = notif.opened_time.date().strftime('%d-%m-%Y') if notif.opened_time else ""
	    ini += 1
	if ini>2:
		wb.save(settings.HOME_FOLDER + '/notification_report_' + todays_date + '.xlsx')
		print 'success'
	else:
		print 'Empty'
except:
	print 'Please provide the starting date and the ending date.'