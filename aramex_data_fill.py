import json
#import pdb
import sys
import os
# sys.path.append("/path/to/project")
import django
try:
    env = sys.argv[1]
except IndexError:
    raise Exception("Please provide settings [local/ staging/ prod]")

os.environ["DJANGO_SETTINGS_MODULE"] = "zapyle_new.settings.{}".format(env)
django.setup()
import pdb
import base64
from django.conf import settings
from openpyxl import load_workbook, Workbook
from zap_apps.logistics.models import AramexStatus

wb = load_workbook(filename=sys.argv[2], data_only=True)
ws = wb['Sheet1']  # ws is now an IterableWorksheet
data = {}

out_wb = Workbook()
out_ws = out_wb.active
# put headers for xcel
out_ws['A1'] = 'Row Num'
out_ws['B1'] = 'Error Desc'
out_ini = 2

# for row in ws.rows:
#     for cell in row:
#         print(cell.value)
# pdb.set_trace()
# for row in ws.rows[1:]:
bulk = []
for ini in xrange(2, ws.max_row+1):
    if ws['G'+str(ini)].value not in ["", None]:
        bulk.append(AramexStatus(status_code=ws['A'+str(ini)].value, problem_code=ws['B'+str(ini)].value, track_code=ws['C'+str(ini)].value, condition_code=ws[
                    'D'+str(ini)].value, code_description=ws['E'+str(ini)].value, customer_description=ws['F'+str(ini)].value, logistic_status=ws['G'+str(ini)].value))
if bulk:
    AramexStatus.objects.bulk_create(bulk)
    print "Success"
else:
    print "Nothing to upload"
