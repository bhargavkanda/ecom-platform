import json
# import pdb
import sys, os
# sys.path.append("/path/to/project")
import django

try:
    env = sys.argv[1]
except IndexError:
    raise Exception("Please provide settings [local/ staging/ prod]")

os.environ["DJANGO_SETTINGS_MODULE"] = "zapyle_new.settings.{}".format(env)
django.setup()
import pdb
import base64
from django.conf import settings
from openpyxl import load_workbook, Workbook
from zap_apps.zap_catalogue.models import Hashtag
from zap_apps.zap_catalogue.product_serializer import ZapProductSerializer, ProductImageSerializer, NumberOfProductSrlzr
from zap_apps.zapuser.models import ZapExclusiveUserData
import urllib2
import re

wb = load_workbook(filename=sys.argv[2], data_only=True)
ws = wb['Sheet1']  # ws is now an IterableWorksheet
data = {}

out_wb = Workbook()
out_ws = out_wb.active
# put headers for xcel
out_ws['A1'] = 'Row Num'
out_ws['B1'] = 'Error Desc'
out_ini = 2

# for row in ws.rows:
#     for cell in row:
#         print(cell.value)
# pdb.set_trace()
# for row in ws.rows[1:]:
for ini in xrange(2, ws.max_row + 1):
    try:
        # pdb.set_trace()
        if not ws['A' + str(ini)].value:
            break
        elif ws['A' + str(ini)].value == 'new_size':
            if saved:

                quantity = int(ws['AC' + str(ini)].value) or 1
                no_pro_data = {'size': int(ws['Z' + str(ini)].value), 'product': product.id, 'quantity': quantity,
                               'time_to_make': str(ws['AI' + str(ini)].value), 'alternate_size': int(ws['AB' + str(ini)].value)}
                no_pro_srlzr = NumberOfProductSrlzr(data=no_pro_data)
                if no_pro_srlzr.is_valid():
                    no_pro_srlzr.save()
            else:
                out_ws['A' + str(out_ini)] = ini
                out_ws['B' + str(out_ini)] = 'Different size for the above product'
                out_ini += 1
        else:
            data['original_price'] = float(ws['E' + str(ini)].value)
            data['listing_price'] = float(ws['F' + str(ini)].value)
            if data['original_price'] >= data['listing_price']:
                images_list = re.split(', |,', ws['B' + str(ini)].value)
                img_ids = []
                for img in images_list:
                    if img.strip():
                        try:
                            req = urllib2.Request(img, headers={'User-Agent': "Magic Browser"})
                            con = urllib2.urlopen(req)
                            encoded = base64.b64encode(con.read())
                            img_serializer = ProductImageSerializer(
                                data={'image': encoded})
                            if img_serializer.is_valid():
                                img_data = img_serializer.save()
                                img_ids.append(img_data.id)
                        except:
                            pass
                if img_ids:
                    data['images'] = img_ids
                    data['title'] = ws['C' + str(ini)].value
                    data['description'] = ws['D' + str(ini)].value

                    # pdb.set_trace()
                    data['sale'] = str(int(ws['H' + str(ini)].value))

                    if ws['J' + str(ini)].value:
                        data['age'] = str(int(ws['J' + str(ini)].value))
                    else:
                        data['age'] = None

                    if ws['L' + str(ini)].value:
                        data['condition'] = str(int(ws['L' + str(ini)].value))
                    else:
                        data['condition'] = None

                    if ws['M' + str(ini)].value:
                        data['percentage_commission'] = ws['M' + str(ini)].value

                    if int(ws['O' + str(ini)].value) == 1:
                        data['with_zapyle'] = True
                    data['user'] = int(ws['P' + str(ini)].value)
                    if ws['Q' + str(ini)].value:
                        data['pickup_address'] = int(ws['Q' + str(ini)].value)
                    else:
                        data['pickup_address'] = None
                    # data['user'] = 1551
                    # data['pickup_address'] = 74
                    data['style'] = int(ws['S' + str(ini)].value)
                    data['brand'] = int(ws['U' + str(ini)].value)
                    data['product_category'] = int(ws['W' + str(ini)].value)
                    data['size_type'] = ws['X' + str(ini)].value
                    #

                    # data['size'] = [int(row[25].value)]

                    data['occasion'] = int(ws['AE' + str(ini)].value)

                    data['color'] = int(ws['AG' + str(ini)].value)
                    data['discount'] = (data['original_price'] - data['listing_price']) / data['original_price']
                    data['status'] = '0'
                    data['completed'] = True
                    data['time_to_make'] = ws['AI' + str(ini)].value
                    data['alternate_size'] = ws['AB' + str(ini)].value
                    data['partner_id'] = ws['AJ' + str(ini)].value

                    # if int(data['user']) in [788, 802]:
                    #
                    words = re.findall(r'[#]\w+', data['description'])
                    if words:
                        tags = []
                        for i in words:
                            tags_data = Hashtag.objects.get_or_create(tag=i)
                            tags.append(tags_data[0].id)
                        data['tags'] = tags
                    srlzr = ZapProductSerializer(data=data)
                    if srlzr.is_valid():
                        print 'success'
                        saved = True
                        product = srlzr.save()
                        # for size in data['size']:
                        quantity = int(ws['AC' + str(ini)].value) or 1
                        no_pro_data = {'size': int(ws['Z' + str(ini)].value), 'product': product.id,
                                       'quantity': quantity, 'time_to_make': str(ws['AI' + str(ini)].value),
                                       'alternate_size': int(ws['AB' + str(ini)].value)}
                        no_pro_srlzr = NumberOfProductSrlzr(data=no_pro_data)
                        if no_pro_srlzr.is_valid():
                            no_pro_srlzr.save()
                        if int(data['user']) in [788, 802]:
                            # pdb.set_trace()
                            zap = ZapExclusiveUserData.objects.get(id=int(ws['AF' + str(ini)].value))
                            zap.products.add(product.id)
                    else:
                        print 'failed - invalid serialiser '+str(srlzr.errors)
                        saved = False
                        out_ws['A' + str(out_ini)] = ini
                        out_ws['B' + str(out_ini)] = str(srlzr.errors)
                        out_ini += 1
                else:
                    print 'failed - images corrupt'
                    saved = False
                    out_ws['A' + str(out_ini)] = ini
                    out_ws['B' + str(out_ini)] = "All images are corrupt"
                    out_ini += 1
            else:
                print 'failed - listing price > original price'
                saved = False
                out_ws['A' + str(out_ini)] = ini
                out_ws['B' + str(out_ini)] = "Listing price is greater than origin price"
                out_ini += 1

    except Exception as e:
        print 'exc failed', e
        # pdb.set_trace()
        # print e
        out_ws['A' + str(out_ini)] = ini
        out_ws['B' + str(out_ini)] = str(e)
        out_ini += 1
if out_ini > 2:
    # pdb.set_trace()
    out_wb.save('err_' + sys.argv[2])